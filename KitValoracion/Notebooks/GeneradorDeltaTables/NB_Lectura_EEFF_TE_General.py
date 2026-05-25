# Databricks notebook source
#Hugo
# %pip install --quiet openpyxl
# dbutils.library.restartPython()

#lectura_carpetas = '[["CAPEX", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/CAPEX/Capex_3Q_2024.xlsx", "CapexPorEmpresa", "CAPEX"], ["Colombia", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Colombia/Kit_Valoracion_ITCO_3Q_2024.xlsx", "ITCO (propuesta)", "EEFF_TE_Col"]]'

#lectura_carpetas = '[["CAPEX", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/CAPEX/Capex_3Q_2024.xlsx", "CapexPorEmpresa", "CAPEX"], ["Colombia", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Colombia/Kit_Valoracion_ITCO_3Q_2024.xlsx", "ITCO (propuesta)", "EEFF_TE_Col"], ["CTEEP", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/CTEEP/CTEEP Consolidado_Kit Inversionista vISA CTEEP_3Q_2024_.xlsx", "EEFF_TE_Bra", "EEFF_TE_Bra"], ["Costera", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Costera/EEFF Costera_3Q_2024.xlsx", "EEFF_V\u00edas_Col", "EEFF_V\u00edas_Col"], ["Deuda", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx", "PerfilD", "Perfil deuda"], ["Dividendos Filiales", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Dividendos Filiales/Dividendos Filiales_3Q_2024.xlsx", "Dividendos", "Dividendos"], ["Interchile", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Interchile/CHILE_TE_3Q_2024.xlsx", "EEFF_TE_Chi", "EEFF_TE_Chi"], ["Participaciones Empresas", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Participaciones Empresas/RP_Participacion Directa & Efectiva0924_3Q_2024.xlsm", "Reporte_SP", "Participaci\u00f3n en compa\u00f1\u00edas"], ["Per\u00fa", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Per\u00fa/EEFF TE Per_3Q_2024.xlsx", "EEFF_TE_Per", "EEFF_TE_Per"], ["Proyectos", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Proyectos/Informaci\u00f3n Negocio TE_3Q_2024.xlsx", "Construcci\u00f3n proyetos", "Proyectos"], ["Transelca", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Transelca/KIT_Valoracion_Transelca_3Q_2024.xlsx", "EEFF_TE_Col", "EEFF_TE_Col"], ["Trafico v\u00edas", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Trafico v\u00edas/InfoV\u00edas-ReportesTrimestralesISA-paraRI_3Q_2024.xlsx", "Tr\u00e1fico v\u00edas", "Tr\u00e1fico v\u00edas"], ["V\u00edas Chile", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/V\u00edas Chile/EEFF V\u00edas Chile_3Q_2024.xlsx", "EEFF_V\u00edas_Chi", "EEFF_V\u00edas_Chi"], ["Vencimiento Concesiones Viales", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Vencimiento Concesiones Viales/Vencimiento Concesiones_3Q_2024.xlsx", "Vencimiento Concesiones", "Vencimiento Concesiones"], ["Deuda", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx", "KitCreditos", "Deuda consolidada cr\u00e9ditos"], ["Deuda", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx", "KitBonos", "Deuda consolidada bonos"]]'

# full_ano = '2024'
# trimestre = '3'
# catalogo = 'kitvaloracion'
# esquema = 'base'

# COMMAND ----------

import pandas as pd
import re
import json
import unicodedata
from pyspark.sql import functions as F
from functools import reduce
from pyspark.sql.window import Window
from pyspark.sql.functions import col, when, split, lit, regexp_replace, lower,udf
from datetime import datetime 
from openpyxl import load_workbook
from delta.tables import DeltaTable
from pyspark.sql.types import StringType

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
catalogo = dbutils.widgets.get('CatalogoDelta')
esquema = dbutils.widgets.get('EsquemaDelta')

ano = int(str(full_ano)[2:])
ruta = f'/Volumes/{catalogo}/{esquema}/tbl_eeff'
table_name = f"{catalogo}.{esquema}.tbl_eeff"

df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn('Carpeta', normalizar_udf(col('Carpeta'))) \
    .filter(
        (col('Carpeta') == 'colombia') |
        (col('Carpeta') == 'costera') |
        (col('Carpeta') == 'cteep') |
        (col('Carpeta') == 'interchile') |
        (col('Carpeta') == 'peru') |
        (col('Carpeta') == 'transelca') |
        (col('Carpeta') == 'vias chile') 
    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]

# COMMAND ----------

try:
    for i in range(len(rutaArchivos)):
        print(f'Archivo {i+1}/{len(rutaArchivos)}: {rutaArchivos[i]}...')
        print(f'Hoja: {hojas_origen[i]}...')
        print('=' * 80)

        source_wb = load_workbook(rutaArchivos[i][5:], data_only=True)
        source_ws = source_wb[hojas_origen[i]]
        encontrado = False

        datos_encontrado = False
        for fila in range(1, source_ws.max_row + 1):
            for columna in range(1, source_ws.max_column + 1):
                valor = str(source_ws.cell(row=fila, column=columna).value).lower()
                                
                if 'estados de resultados' in valor:
                    fila_inicio_datos = fila - 1
                    columna_inicio_datos = columna -1
                    datos_encontrado = True
                    break

            if datos_encontrado:
                #print(f'[{fila}.{columna}] = {valor}')
                break

        encabezado_encontrado = False
        for fila in range(1, source_ws.max_row + 1):
            for columna in range(1, source_ws.max_column + 1):
                valor = str(source_ws.cell(row=fila, column=columna).value)
                
                if valor == f'{trimestre}T{ano}' or valor == f'{trimestre}Q{ano}':
                    fila_encabezado = fila - 1
                    columna_encabezado = columna - 1
                    nombre_columna = valor
                    encabezado_encontrado = True
                    break

            if encabezado_encontrado:
                #print(f'[{fila}.{columna}] = {valor}')
                break
        
        df_pandas = pd.read_excel(
            rutaArchivos[i][5:],
            sheet_name=hojas_origen[i],
            engine='openpyxl',
            header=None
        )
        df_pandas = df_pandas.astype(str)

        fila_datos_encabezado = df_pandas.iloc[fila_encabezado,columna_inicio_datos:]
        columna_descripciones = df_pandas.iloc[:,columna_inicio_datos]

        df_convert = df_pandas.iloc[:,columna_inicio_datos:]
        df_convert.columns = fila_datos_encabezado

        columna_descripcion = df_convert.iloc[:, 0]
        columna_valores = df_convert[f"{nombre_columna}"]
        df_subset = pd.concat([columna_descripcion, columna_valores], axis=1)

        df = spark.createDataFrame(df_subset)
        
        df_inicial = df\
            .withColumnRenamed(df.columns[0], 'Descripcion')\
            .withColumnRenamed(df.columns[1], 'Valor')\
            .replace('nan', None)

        df_empresas = df_inicial\
        .filter(
            (col('Descripcion').like('ESTADOS SEPARADOS%')) |
            (col('Descripcion').like('ESTADOS DE RESULTADOS SEPARADOS%')) |
            (col('Descripcion').like('ESTADOS DE RESULTADOS CONSOLIDADO%')) |
            (col('Descripcion').like('ESTADOS CONSOLIDADOS DE SITUACIÓN%')) 
        )\
        .select(col('Descripcion'))\
        .withColumnRenamed("Descripcion", "TipoEstado")

        df_limpio = df_inicial\
        .filter(
            (~col('Descripcion').like('Valores expresados en%')) & 
            (~col('Descripcion').like('Cifras expresadas en%')) & 
            (col('Descripcion').like('Total%') != True) &
            (col('Descripcion').like('TOTAL%') != True) &
            (col('Descripcion').like('IFRS Results') != True)
        )\
        .withColumn("Descripcion", 
                    F.regexp_replace(
                        F.col("Descripcion"), "PASIVO MAS PATRIMONIO DE LOS ACCIONISTAS", "PASIVO"
                        )
                    )\
        .withColumn("Descripcion", 
                    F.regexp_replace(
                        F.col("Descripcion"), "Liabilities and Shareholders' Equity", "Liabilities"
                        )
                    )\
        .withColumn("Descripcion", 
                    F.regexp_replace(
                        F.col("Descripcion"), "SHAREHOLDERS' EQUITY", "EQUITY"
                        )
                    )\
        .withColumn("Descripcion", 
                    F.regexp_replace(
                        F.col("Descripcion"), "Patrimonio de los accionistas", "Patrimonio"
                        )
                    )\
        .withColumn("Descripcion", 
                    F.regexp_replace(
                        F.col("Descripcion"), "Total patrimonio de los accionistas", "Total Patrimonio"
                        )
                    )\
        .withColumn("Descripcion", 
                    F.regexp_replace(
                        F.col("Descripcion"), "TOTAL SHAREHOLDER'S EQUITY", "Total Equity"
                        )
                    )\
        .withColumn("Descripcion", F.regexp_replace(F.col("Descripcion"), "Assets \\(BRL thousand\\)", "Assets") )

        #\.withColumn("Descripcion", F.regexp_replace(F.col("Descripcion"), r"(?i)\bGross(?:\s+\w+)*\b","Gross") )

        # 1) Conviertes df a un DF "ordenado" (asignando row_id para preservar el orden de las filas).
        w_all = Window.orderBy(F.monotonically_increasing_id())
        df_window = df_limpio.withColumn("row_id", F.row_number().over(w_all))

        # 2) Obtenemos la lista de empresas desde df_empresas (si es pequeño) o la manejamos de otra forma.
        empresas_list = [row["TipoEstado"] for row in df_empresas.collect()]

        # 3) Creamos una columna "empresa_marker" que copia la Descripcion si es una empresa, o None si no lo es.
        df_mark = df_window\
            .withColumn(
                "empresa_marker",
                F.when(F.col("Descripcion").isin(empresas_list), F.col("Descripcion"))
            )

        # 4) Definimos una ventana que ordene por row_id y va desde el inicio hasta la fila actual.
        #    Luego aplicamos last(..., ignorenulls=True) para "arrastrar" el último valor de empresa_marker.
        w = Window.orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

        df_filled = df_mark\
            .withColumn(
                "Empresa",
                F.last("empresa_marker", ignorenulls=True).over(w)
            )

        df_filled = df_filled\
            .filter(
                col("Empresa") != col('Descripcion')
            )\
            .select(col("Empresa"),col("Descripcion"),col("Valor"))
        
        # 1) Definir una lista de tipos contables que deseas detectar
        tipos_especiales = ["ACTIVO","ASSETS", "PASIVO","LIABILITIES", "PATRIMONIO","EQUITY","GROSS" ]
        # (Para 'Utilidad' no se hace un match directo, sino que la usaremos como "por defecto")

        # 2) Crear una columna "tipo_marker" que solo se llena si Descripcion es uno de los tipos
        df_mark = df_filled.withColumn(
            "tipo_marker",
            F.when(F.upper(F.col("Descripcion")).isin(tipos_especiales), F.initcap(F.col("Descripcion")))
            .otherwise(None)
        )

        # Asegurarse de que la columna 'row_id' existe
        df_mark = df_mark.withColumn("row_id", F.monotonically_increasing_id())

        w = Window.partitionBy("Empresa").orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

        df_filled = df_mark.withColumn(     
            "tipo_filled",
            F.last("tipo_marker", ignorenulls=True).over(w)
        )

        # 4) Si "tipo_filled" es nulo, significa que todavía no hemos visto "ACTIVO"/"PASIVO"/"PATRIMONIO"
        #    => asignamos "Utilidad" como valor por defecto
        df_final = df_filled.withColumn(
            "Tipo",
            F.coalesce(F.col("tipo_filled"), F.lit("Utilidad"))
        )

        # 5) Ver el resultado ordenado por row_id
        df_final = df_final\
            .select(col('row_id'),col('Descripcion'),col('Empresa'),col('Valor'), col('Tipo').alias('ValorContable'))\
            .orderBy("row_id")
        
        # 1) Definir una lista de tipos contables que deseas detectar
        tipos_especiales = ["Activo corriente","Activo no corriente", "Pasivo corriente","Pasivo no corriente", "CURRENT","NON-CURRENT","CURRENT ASSETS","NON-CURRENT ASSETS","CURRENT LIABILITIES","NON-CURRENT LIABILITIES" ]
        # (Para 'Utilidad' no se hace un match directo, sino que la usaremos como "por defecto")

        # 2) Crear una columna "tipo_marker" que solo se llena si Descripcion es uno de los tipos
        df_marcado = df_final.withColumn(
            "corriente_tipo",
            F.when(F.col("Descripcion").isin(tipos_especiales),  F.initcap(F.col("Descripcion")))
            .otherwise(None)
        )

        # Asegurarse de que la columna 'row_id' existe
        df_marcado = df_marcado.withColumn("row_id", F.monotonically_increasing_id())

        w = Window.partitionBy("Empresa").orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

        df_relleno = df_marcado.withColumn(     
            "relleno_tipo",
            F.last("corriente_tipo", ignorenulls=True).over(w)
        )

        # 4) Si "tipo_filled" es nulo, significa que todavía no hemos visto "ACTIVO"/"PASIVO"/"PATRIMONIO"
        #    => asignamos "Utilidad" como valor por defecto
        df_ultimo = df_relleno\
            .withColumn(
                "TipoCorriente",
                when(
                    (
                        (col("ValorContable") == 'Activo') |
                        (col("ValorContable") == 'Pasivo') |
                        (col("ValorContable") == 'Assets') |
                        (col("ValorContable") == 'Liabilities')
                    ),
                    F.coalesce(F.col("relleno_tipo"))
                )\
                .otherwise("N/A")
            
            )\
            .select(col('row_id'),col('Descripcion'),col('Empresa'),col('Valor'),col('ValorContable'), col('TipoCorriente'))
            
        fecha = datetime.now()

        #Eliminar las filas con información innecesarias

        df_limpieza = df_ultimo\
        .withColumn(
            "TipoCorriente",
            F.when(
                (
                    (
                        (col('ValorContable').like('Assets')) &
                        (col('TipoCorriente').like('Current'))
                    ) |
                    (
                        (col('ValorContable').like('Assets')) &
                        (col('TipoCorriente').like('Current Assets'))
                    ) 
                ),
                "Activo Corriente"
            )\
            .otherwise(
                F.when(
                    (
                        (
                            (col('ValorContable').like('Liabilities')) &
                            (col('TipoCorriente').like('Current'))
                        ) |
                        (
                            (col('ValorContable').like('Liabilities')) &
                            (col('TipoCorriente').like('Current Liabilities'))
                        ) 
                    ),
                    "Pasivo Corriente"
                )\
                .otherwise(
                    F.when(
                        (
                            (
                                (col('ValorContable').like('Assets')) &
                                (col('TipoCorriente').like('Non-current'))
                            ) |
                            (
                                (col('ValorContable').like('Assets')) &
                                (col('TipoCorriente').like('Non-current Assets'))
                            ) 
                        ),
                        "Activo No Corriente"
                    )\
                    .otherwise(
                        F.when(
                            (
                                (
                                    (col('ValorContable').like('Liabilities')) &
                                    (col('TipoCorriente').like('Non-current'))
                                ) |
                                (
                                    (col('ValorContable').like('Liabilities')) &
                                    (col('TipoCorriente').like('Non-current Liabilities'))
                                ) 
                            ),
                            "Pasivo No Corriente"
                        )\
                        .otherwise(col('TipoCorriente'))
                    )
                )
            )
        )\
        .filter(
            (col('Descripcion').like('%ACTIVO%') != True) &
            (col('Descripcion').like('%Activo corriente%') != True) &
            (col('Descripcion').like('%Activo no corriente%') != True) &
            (col('Descripcion').like('%PASIVO%') != True) &
            (col('Descripcion').like('%Pasivo corriente%') != True) &
            (col('Descripcion').like('%Pasivo no corriente%') != True) &
            (col('Descripcion').like('%Patrimonio%') != True) &
            (col('Descripcion').like('%Assets%') != True) &
            (col('Descripcion').like('%CURRENT%') != True) &
            (F.upper(col('Descripcion')) != F.upper(col('Empresa'))) &
            (F.upper(col('Descripcion')) != F.upper(col('ValorContable'))) &
            (~col('Descripcion').like('%(BRL thousand)%')) &
            (col('Descripcion').like('%EQUITY%') != True) 
        )\
        .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Gross", "Utilidad") )\
        .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Assets", "Activo") )\
        .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Liabilities", "Pasivo") )\
        .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Equity", "Patrimonio") )\
        .filter(~col('Descripcion').isin(empresas_list))\
        .withColumn('Año', lit(full_ano))\
        .withColumn('Trimestre', lit(trimestre))\
        .withColumn("FECHA_PUBLICACION",lit(fecha))\
        .withColumn("Valor",col("Valor").cast("double"))\
        .fillna(0)\
        .withColumn(
            "Estado",
            F.trim(
                    F.split(F.col("Empresa"), "-")[0]
                )
        )\
        .withColumn(
            "Empresa",
            F.trim(
                    F.split(F.col("Empresa"), "-")[1]
            )
        )\
        .select(
            col('Estado'),
            col('Empresa'),
            col('Descripcion'),
            col('ValorContable'),
            col('TipoCorriente'),
            col('valor'),
            col('Año'),
            col('Trimestre'),
            col('FECHA_PUBLICACION')
        )

        if i == 0:
            df_guardado = spark.createDataFrame([], df_limpieza.schema)

        df_guardado = df_guardado.union(df_limpieza)

    # Not supported in serverless, comment out
    # spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

    if not spark.catalog.tableExists(table_name):
        df_guardado.write \
            .format("delta") \
            .mode('overwrite') \
            .partitionBy("`Año`","Trimestre","Empresa","Estado","ValorContable","TipoCorriente","Descripcion") \
            .saveAsTable(table_name)
    else:
        df_guardado.createOrReplaceTempView("tmp_eeff")
        spark.sql(
        f"""
            MERGE INTO {table_name} AS target
                USING tmp_eeff AS source
                ON  target.`Año`       = source.`Año`
                AND target.Trimestre   = source.Trimestre
                AND target.Empresa       = source.Empresa
                AND target.Estado    = source.Estado
                AND target.ValorContable    = source.ValorContable
                AND target.TipoCorriente    = source.TipoCorriente
                AND target.Descripcion    = source.Descripcion

                WHEN MATCHED THEN
                UPDATE SET *
                WHEN NOT MATCHED THEN
                INSERT *;
        """
    )
except Exception as ex:
    print('Task Fail--> ### ' + str(ex) + ' ###')

# COMMAND ----------

#display(spark.sql("SELECT COUNT(*) AS num_filas FROM kitvaloracion.base.tbl_eeff"))
#display(spark.sql("SELECT * FROM kitvaloracion.base.tbl_eeff LIMIT 3"))
