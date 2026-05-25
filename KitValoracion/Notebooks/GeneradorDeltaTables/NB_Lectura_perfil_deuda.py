# Databricks notebook source
import pyspark.sql.functions as F
import unicodedata
import pandas as pd
import json
from pyspark.sql.functions import col, lit, when, regexp_replace, lower, udf
from pyspark.sql.window import Window
from datetime import datetime 
from openpyxl import load_workbook
from pyspark.sql.types import *
from delta.tables import DeltaTable 


# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

# rutaArchivos = ['abfs:/Volumes/test_data/base/4q23/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx']
# hojas_origen = ['PerfilD']
# ano = 2026

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
catalogo = dbutils.widgets.get('CatalogoDelta')
esquema = dbutils.widgets.get('EsquemaDelta')

#ruta = f'/Volumes/{catalogo}/{esquema}/'
table_name = f"{catalogo}.{esquema}.tbl_perfil_deuda"

df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .filter(
        (normalizar_udf(col('Carpeta')) == 'deuda') &
        (normalizar_udf(col('HojaOrigen')) == 'perfild') 

    )

display(df_carpetas)

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]

# COMMAND ----------

source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
source_ws = source_wb[hojas_origen[0]]
encontrado = False

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        
        valor = str(source_ws.cell(row=fila, column=columna).value).upper()
        #print(f'{fila}.{columna}) {valor}')

        if 'PERFIL DE VENCIMIENTO POR AÑOS' in valor:
            fila_encabezado = fila - 1
            columna_inicio = columna - 1
            encontrado = True
            break

    if encontrado:
        break

print(f'fila_encabezado: {fila_encabezado}, columna_inicio: {columna_inicio}')

# COMMAND ----------

lista_color_celda = []
#i = 1
for fila in range(fila_encabezado + 2, source_ws.max_row + 1):
    celda = source_ws.cell(row=fila, column=columna_inicio + 2)
    #print(f'{i}) {celda.value} --> {celda.fill.fgColor.tint}')
    #i = i +1
    lista_color_celda.append(celda.fill.fgColor.tint)

# COMMAND ----------

df_pandas = pd.read_excel(
    rutaArchivos[0][5:],
    sheet_name= hojas_origen[0],
    engine='openpyxl',
    header=fila_encabezado
    )
df_pandas = df_pandas\
            .astype(str)

print(len(df_pandas.iloc[:fila_encabezado,:columna_inicio]))
display(df_pandas)

# COMMAND ----------

#lista_color_celda = lista_color_celda[:len(df_pandas.iloc[:fila_encabezado,:columna_inicio]) - len(lista_color_celda)]
#len(lista_color_celda)

# COMMAND ----------

df_colores = spark.createDataFrame(lista_color_celda,'double')

display(df_colores)

# COMMAND ----------

df = spark.createDataFrame(df_pandas)


display(df)

# COMMAND ----------

df_inicial = df\
    .replace('nan', None)\
    .select(
        col(df.columns[1]).alias('Compañia'),
        col(df.columns[2]).alias('Moneda'),
        col(f'{ano}').alias('Valor')
    )

display(df_inicial)

# COMMAND ----------

ventana = Window.orderBy(F.lit(1))
df_inicial_indexado = df_inicial.withColumn("row_id", F.row_number().over(ventana))
df_colores_indexado = df_colores.withColumn("row_id", F.row_number().over(ventana))

# Realizamos el join sobre el índice y luego eliminamos la columna row_id
df_unido = df_inicial_indexado.join(df_colores_indexado, on="row_id").drop("row_id")

df_unido = df_unido\
    .filter(
        ~(
            (col('Compañia').isNull()) &
            (col('Moneda').isNull()) &
            (col('Valor').isNull())
        )
    )

display(df_unido)

# COMMAND ----------

df_relleno = df_unido.withColumn("row_id", F.monotonically_increasing_id())

ventana_contexto_relleno = (
    Window
    .orderBy("row_id")
    .rowsBetween(1, Window.unboundedFollowing)  # de la fila siguiente hasta la última
)

df_relleno = df_relleno.withColumn(
    "Pais",
    F.when(
        F.col("value") != 0,            # Si la fila actual no es cero
        F.col("Compañia")                  # usamos el propio valor
    ).otherwise(
        # Si la fila actual es cero, buscamos en las filas siguientes:
        F.first(                        # tomamos el primer Valor no-cero
            F.when(F.col("Value") != 0, F.col("Compañia")),  # si es 0 => null
            ignorenulls=True
        ).over(ventana_contexto_relleno)
    )
)

display(df_relleno)

# COMMAND ----------

ventana_valor_anterior = Window.orderBy("row_id")
fecha = datetime.now()

df_tratado = df_relleno\
    .filter(
        (col('Pais').isNotNull()) &
        (col('Pais') != 'TOTAL DEUDA')
    )\
    .withColumn(
        "Compañia",
        F.when(
            F.col("Compañia").isNotNull(),            # Si la fila actual no es cero
            F.col("Compañia")                  # usamos el propio valor
        ).otherwise(
             F.lag("Compañia", 1).over(ventana_valor_anterior)
            
        )
    )\
    .withColumn(
        'Año',
        lit(f'{ano}')
    )\
    .withColumn(
        'trimestreEvaluado',
        lit(f'{trimestre}')
    )\
    .withColumn(
        'FECHAPUBLICACION',
        lit(fecha.strftime('%Y-%m-%d'))
    )

df_final = df_tratado\
    .select(
        col('Pais').alias('Pais'),
        col('Compañia').alias('Compania'),
        col('Moneda'),
        col('Valor').cast('double'),
        col('Año').alias('Ano'),
        col('trimestreEvaluado').alias('trimestre'),
        col('FECHAPUBLICACION'),
    )
    
display(df_final)

# COMMAND ----------

try:
    # Not supported in serverless, comment out
    # spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

    if not spark.catalog.tableExists(table_name):
        df_final.write \
            .format("delta") \
            .mode('overwrite') \
            .partitionBy("Ano","trimestre","Moneda","Pais","Compania")  \
            .saveAsTable(table_name)
    else:
        df_final.createOrReplaceTempView("tmp_perfil_deuda")
        spark.sql(
            f"""
                MERGE INTO {table_name} AS target
                    USING tmp_perfil_deuda AS source
                    ON  target.Ano       = source.Ano
                    AND target.Trimestre   = source.Trimestre
                    AND target.Moneda       = source.Moneda
                    AND target.Pais    = source.Pais
                    AND target.Compania    = source.Compania

                    WHEN MATCHED THEN
                    UPDATE SET *
                    WHEN NOT MATCHED THEN
                    INSERT *;
            """
        )
except Exception as ex:
    print('Task Fail--> ### ' + str(ex) + ' ###')        
