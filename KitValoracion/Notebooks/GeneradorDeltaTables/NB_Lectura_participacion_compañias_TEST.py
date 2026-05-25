# Databricks notebook source
import pandas as pd
import pyspark.sql.functions as F
import json
from openpyxl import load_workbook
from pyspark.sql.functions import col, lit, when,regexp_replace
from pyspark.sql.window import Window
from datetime import datetime
from delta.tables import DeltaTable

# COMMAND ----------

rutaArchivos = ['abfs:/Volumes/test_data/base/4q24/Participaciones Empresas/RP_Participacion Directa & Efectiva1224.xlsm']
hojas_origen = ['Reporte_SP']

# COMMAND ----------

# Cargar los libros de Excel
source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)

# Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
source_ws = source_wb[hojas_origen[0]]

# COMMAND ----------

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, 5):
        valor = str(source_ws.cell(row=fila, column=columna).value).lower()
        #print(f'{fila}.{columna}) {valor}')
        if valor == 'tipo de \ncontrol' :
            fila_encabezado = fila - 1
            columna_inicio = columna - 1
print(fila_encabezado, columna_inicio)

# COMMAND ----------

df_pandas = pd.read_excel(
    rutaArchivos[0][5:],
    sheet_name=hojas_origen[0],
    engine='openpyxl',
    header=fila_encabezado
    )
df_pandas = df_pandas.iloc[:,columna_inicio:]

# COMMAND ----------

df_pandas = df_pandas\
            .astype(str)
        
df = spark.createDataFrame(df_pandas)

df = df\
    .select(
        col('Tipo de \nControl').alias('TipoDeControl'), col('Negocio'), col('País'), col('SociedadGL'), 
        col('% Directo ISA').alias('PorcentajeDirectoISA'), col('% Indirecto a través de Otras Filiales').alias('PorcentajeIndirectoOtrasFiliales'), 
        col('% Efectivo ISA').alias('PorcentajeEfectivoISA'), col('SocIndirecta'))\
    .replace('nan', None)

display(df)

# COMMAND ----------

# Agregar un índice para mantener el orden (puedes ajustar este método según tu fuente de datos)
df = df.withColumn("id_fila", F.monotonically_increasing_id())

# Definir una ventana ordenada por 'id'
ventana_orden = Window.orderBy("id_fila")

# Definir una ventana que ordene por el índice y que abarque desde la primera fila hasta la actual
ventana_entre_datos = Window.orderBy("id_fila").rowsBetween(Window.unboundedPreceding, 0)


df_ultimo = df\
    .withColumn("SociedadGL", F.last("SociedadGL", ignorenulls=True).over(ventana_entre_datos))\
    .withColumn("País", F.last("País", ignorenulls=True).over(ventana_entre_datos))\
    .withColumn("Negocio", F.last("Negocio", ignorenulls=True).over(ventana_entre_datos))\
    .withColumn("TipoDeControl", F.last("TipoDeControl", ignorenulls=True).over(ventana_entre_datos))\
    .withColumn(
        "PorcentajeDirectoISA", 
        when(
            (
                col("SociedadGL").isNull() &
                F.lag("PorcentajeDirectoISA", 1).over(ventana_orden).isNotNull() 
            ),
            F.last("PorcentajeDirectoISA", ignorenulls=True).over(ventana_entre_datos)
        )\
        .otherwise(col('PorcentajeDirectoISA'))
        
    )

windowSpec = (
    Window.partitionBy("SociedadGL")
          .orderBy("id_fila")
          .rowsBetween(Window.unboundedPreceding, Window.unboundedFollowing)
)
columns_to_fill = [c for c in df_ultimo.columns if c not in ("SociedadGL", "id_fila")]

for c in columns_to_fill:
    # 4a. Creamos una columna temporal con el valor de la primera fila del grupo
    first_col = c + "_first"
    df_ultimo = df_ultimo.withColumn(first_col, F.first(c, ignorenulls=False).over(windowSpec))
    
    # 4b. Rellenamos la columna original con ese valor si está en null
    df_ultimo = df_ultimo.withColumn(c, F.coalesce(F.col(c), F.col(first_col)))
    

df_limpio = df_ultimo\
    .select(
    col('TipoDeControl'),
    col('Negocio'),
    col('País').alias('Pais'),
    col('SociedadGL'),
    col('PorcentajeDirectoISA'),
    col('PorcentajeIndirectoOtrasFiliales'),
    col('PorcentajeEfectivoISA'),
    col('SocIndirecta')
)

display(df_limpio)
    

# COMMAND ----------

fecha = datetime.now()
    
df_final = df_limpio\
    .withColumn("PorcentajeDirectoISA",F.concat_ws(" ", F.col("PorcentajeDirectoISA") * 100,lit("%")))\
    .withColumn("PorcentajeIndirectoOtrasFiliales",F.concat_ws(" ", F.col("PorcentajeIndirectoOtrasFiliales") * 100,lit("%")))\
    .withColumn("PorcentajeEfectivoISA",F.concat_ws(" ", F.col("PorcentajeEfectivoISA") * 100,lit("%")))\

for c in df_final.columns:
    df_final = df_final.withColumn(
        c, 
        F.when(
            F.col(c).cast("string").contains("#Error"), 
            F.lit("Data Corrupa")
        ).otherwise(F.col(c))
    )
df_final = df_final.withColumn("FECHAPUBLICACION", lit(fecha))

display(df_final)


# COMMAND ----------

spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

table_name = "test_data.base.participacionempresas"

df_final.write \
    .format("delta") \
    .mode("overwrite") \
    .partitionBy("TipoDeControl","Negocio","Pais") \
    .save("/Volumes/test_data/base/deltatables/participacionempresas/")

if not spark.catalog.tableExists(table_name):

    spark.sql(
        """
            CREATE TABLE test_data.base.participacionempresas
            AS SELECT *
            FROM delta.`/Volumes/test_data/base/deltatables/participacionempresas/`
        """
    )
else:
    # Cargar la tabla destino desde Unity Catalog
    tabla_destino = DeltaTable.forName(spark, "test_data.base.participacionempresas")
    
    condicion = """
    target.TipoDeControl = source.TipoDeControl AND 
    target.Negocio = source.Negocio AND 
    target.Pais = source.Pais AND
    target.SociedadGL = source.SociedadGL AND
    target.SocIndirecta = source.SocIndirecta
    """

    tabla_destino.alias("target").merge(
        source = df_final.alias("source"),
        condition = condicion
    ).whenMatchedUpdateAll() \
    .whenNotMatchedInsertAll() \
    .execute()
