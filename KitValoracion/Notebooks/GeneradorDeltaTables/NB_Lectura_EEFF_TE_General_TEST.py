# Databricks notebook source
import pandas as pd
from pyspark.sql import functions as F
from functools import reduce
from pyspark.sql.window import Window
from pyspark.sql.functions import col, when, split, lit
from datetime import datetime 
from openpyxl import load_workbook
import json
from delta.tables import DeltaTable
import re

# COMMAND ----------

#Colombia
# rutaArchivos = ['abfs:/Volumes/test_data/base/4q24/Colombia/Kit_Valoracion_Q4_ITCO.xlsx']
# hojas_origen = ['ITCO (propuesta)']
# trimestre = 4
# full_ano = 2024
# ano = int(str(full_ano)[2:])
# archivo = 'Col_ITCO'

# COMMAND ----------

#Costera
# rutaArchivos = ['absf:/Volumes/test_data/base/4q24/Costera/EEFF Costera.xlsx']
# hojas_origen = ['EEFF_Vías_Col']
# trimestre = 4
# full_ano = 2024
# ano = int(str(full_ano)[2:])
# archivo = 'Costera'

# COMMAND ----------

#CTEEP

# rutaArchivos = ['asbf:/Volumes/test_data/base/4q24/CTEEP/2024/CTEEP Consolidado_Kit Inversionista vISA ENERGIA BRASIL 4Q24_.xlsx']
# hojas_origen = ['EEFF_TE_Bra']
# trimestre = 4
# full_ano = 2024
# ano = int(str(full_ano)[2:])
# archivo = 'CTEEP'



# COMMAND ----------

#Interchile

# rutaArchivos = ['absf:/Volumes/test_data/base/4q23/Interchile/CHILE_TE_2T23.xlsx']
# hojas_origen = ['EEFF_TE_Chi']
# trimestre = 4
# full_ano = 2024
# ano = int(str(full_ano)[2:])
# archivo = 'InterChile'

# COMMAND ----------

#TE Per

# rutaArchivos = ['absf:/Volumes/test_data/base/4q24/Perú/EEFF TE Per 2023 T3.xlsx']
# hojas_origen = ['EEFF_TE_Per']
# trimestre = 4
# full_ano = 2024
# ano = int(str(full_ano)[2:])
# archivo = 'Per'



# COMMAND ----------

#Transelca

# rutaArchivos = ['abfs:/Volumes/test_data/base/4q23/Transelca/KIT_Valoracion_Q4_Transelca 2023.xlsx']
# hojas_origen = ['EEFF_TE_Col']
# trimestre = 4
# full_ano = 2024
# ano = int(str(full_ano)[2:])
# archivo = 'Transelca'




# COMMAND ----------

#Vías Chile

rutaArchivos = ['abfs:/Volumes/test_data/base/4q23/Vías Chile/EEFF Vías Chile 4T2023.xlsx']
hojas_origen = ['EEFF_Vías_Chi']
trimestre = 4
full_ano = 2023
ano = int(str(full_ano)[2:])
archivo = 'Vías_Chile'

# COMMAND ----------

source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
source_ws = source_wb[hojas_origen[0]]


# COMMAND ----------

datos_encontrado = False
for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        valor = str(source_ws.cell(row=fila, column=columna).value).lower()
        #print(f'{fila}.{columna}) {valor}')
        if 'estados de resultados' in valor:
            fila_inicio_datos = fila - 1
            columna_inicio_datos = columna -1
            datos_encontrado = True
            break

    if datos_encontrado:
        break
print(f'fila_inicio_informacion: {fila_inicio_datos}, columna_inicio_informacion: {columna_inicio_datos}')

# COMMAND ----------

encabezado_encontrado = False
for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        valor = str(source_ws.cell(row=fila, column=columna).value)
        #print(f'{fila}.{columna}) {valor}')

        #if bool(re.match(r"^[1-4][T,Q]\d{2}$",valor)):
        if valor == f'{trimestre}T{ano}'\
        or valor == f'{trimestre}Q{ano}':
            fila_encabezado = fila - 1
            columna_encabezado = columna - 1
            nombre_columna = valor
            encabezado_encontrado = True
            break

    if encabezado_encontrado:
        break

print(f'fila_encabezado: {fila_encabezado}, columna_encabezado: {columna_encabezado}')

# COMMAND ----------

pd.options.display.max_rows = None
pd.options.display.max_columns = None

# COMMAND ----------

df_pandas = pd.read_excel(
    rutaArchivos[0][5:],
    sheet_name=hojas_origen[0],
    engine='openpyxl',
    header=None
)

df_pandas = df_pandas\
    .astype(str)\

display(df_pandas.iloc[:8,columna_inicio_datos:])

# COMMAND ----------

fila_datos_encabezado = df_pandas.iloc[fila_encabezado,columna_inicio_datos:]
columna_descripciones = df_pandas.iloc[:,columna_inicio_datos]

print(fila_datos_encabezado)
print(columna_descripciones)


# COMMAND ----------

df_convert = df_pandas.iloc[:,columna_inicio_datos:]
df_convert.columns = fila_datos_encabezado

display(df_convert.iloc[:8,:])

# COMMAND ----------

columna_descripcion = df_convert.iloc[:, 0]
columna_valores = df_convert[f"{nombre_columna}"]
df_subset = pd.concat([columna_descripcion, columna_valores], axis=1)

display(df_subset)


# COMMAND ----------

df = spark.createDataFrame(df_subset)

display(df)

# COMMAND ----------

df_inicial = df\
    .withColumnRenamed(df.columns[0], 'Descripcion')\
    .withColumnRenamed(df.columns[1], 'Valor')\
    .replace('nan', None)


display(df_inicial)

# COMMAND ----------

df_empresas = df_inicial\
    .filter(
        (col('Descripcion').like('ESTADOS SEPARADOS%')) |
        (col('Descripcion').like('ESTADOS DE RESULTADOS SEPARADOS%')) |
        (col('Descripcion').like('ESTADOS DE RESULTADOS CONSOLIDADO%')) |
        (col('Descripcion').like('ESTADOS CONSOLIDADOS DE SITUACIÓN%')) 
    )\
    .select(col('Descripcion'))\
    .withColumnRenamed("Descripcion", "TipoEstado")

display(df_empresas)

# COMMAND ----------

df_limpio = df_inicial\
    .filter(
        (~col('Descripcion').like('Valores expresados en%')) & 
        (~col('Descripcion').like('Cifras expresadas en%')) & 
        (col('Descripcion').like('Total%') != True) &
        (col('Descripcion').like('TOTAL%') != True) &
        (col('Descripcion').like('IFRS Results') != True)
    )\
    .withColumn("Descripcion", 
                F.regexp_replace(
                    F.col("Descripcion"), "PASIVO MAS PATRIMONIO DE LOS ACCIONISTAS", "PASIVO"
                    )
                )\
    .withColumn("Descripcion", 
                F.regexp_replace(
                    F.col("Descripcion"), "Liabilities and Shareholders' Equity", "Liabilities"
                    )
                )\
    .withColumn("Descripcion", 
                F.regexp_replace(
                    F.col("Descripcion"), "SHAREHOLDERS' EQUITY", "EQUITY"
                    )
                )\
    .withColumn("Descripcion", 
                F.regexp_replace(
                    F.col("Descripcion"), "Patrimonio de los accionistas", "Patrimonio"
                    )
                )\
    .withColumn("Descripcion", 
                F.regexp_replace(
                    F.col("Descripcion"), "Total patrimonio de los accionistas", "Total Patrimonio"
                    )
                )\
    .withColumn("Descripcion", 
                F.regexp_replace(
                    F.col("Descripcion"), "TOTAL SHAREHOLDER'S EQUITY", "Total Equity"
                    )
                )\
    .withColumn("Descripcion", F.regexp_replace(F.col("Descripcion"), "Assets \\(BRL thousand\\)", "Assets") )

display(df_limpio)

# COMMAND ----------

#\.withColumn("Descripcion", F.regexp_replace(F.col("Descripcion"), r"(?i)\bGross(?:\s+\w+)*\b","Gross") )

# 1) Conviertes df a un DF "ordenado" (asignando row_id para preservar el orden de las filas).
w_all = Window.orderBy(F.monotonically_increasing_id())
df_window = df_limpio.withColumn("row_id", F.row_number().over(w_all))

# 2) Obtenemos la lista de empresas desde df_empresas (si es pequeño) o la manejamos de otra forma.
empresas_list = [row["TipoEstado"] for row in df_empresas.collect()]

# 3) Creamos una columna "empresa_marker" que copia la Descripcion si es una empresa, o None si no lo es.
df_mark = df_window\
    .withColumn(
        "empresa_marker",
        F.when(F.col("Descripcion").isin(empresas_list), F.col("Descripcion"))
    )

# 4) Definimos una ventana que ordene por row_id y va desde el inicio hasta la fila actual.
#    Luego aplicamos last(..., ignorenulls=True) para "arrastrar" el último valor de empresa_marker.
w = Window.orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

df_filled = df_mark\
    .withColumn(
        "Empresa",
        F.last("empresa_marker", ignorenulls=True).over(w)
    )

df_filled = df_filled\
    .filter(
        col("Empresa") != col('Descripcion')
    )\
    .select(col("Empresa"),col("Descripcion"),col("Valor"))

display(df_filled)

# COMMAND ----------

# 1) Definir una lista de tipos contables que deseas detectar
tipos_especiales = ["ACTIVO","ASSETS", "PASIVO","LIABILITIES", "PATRIMONIO","EQUITY","GROSS" ]
# (Para 'Utilidad' no se hace un match directo, sino que la usaremos como "por defecto")

# 2) Crear una columna "tipo_marker" que solo se llena si Descripcion es uno de los tipos
df_mark = df_filled.withColumn(
    "tipo_marker",
    F.when(F.upper(F.col("Descripcion")).isin(tipos_especiales), F.initcap(F.col("Descripcion")))
    .otherwise(None)
)

# Asegurarse de que la columna 'row_id' existe
df_mark = df_mark.withColumn("row_id", F.monotonically_increasing_id())

w = Window.partitionBy("Empresa").orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

df_filled = df_mark.withColumn(     
    "tipo_filled",
    F.last("tipo_marker", ignorenulls=True).over(w)
)

# 4) Si "tipo_filled" es nulo, significa que todavía no hemos visto "ACTIVO"/"PASIVO"/"PATRIMONIO"
#    => asignamos "Utilidad" como valor por defecto
df_final = df_filled.withColumn(
    "Tipo",
    F.coalesce(F.col("tipo_filled"), F.lit("Utilidad"))
)

# 5) Ver el resultado ordenado por row_id
df_final = df_final\
    .select(col('row_id'),col('Descripcion'),col('Empresa'),col('Valor'), col('Tipo').alias('ValorContable'))\
    .orderBy("row_id")

display(df_final)

# COMMAND ----------

# 1) Definir una lista de tipos contables que deseas detectar
tipos_especiales = ["Activo corriente","Activo no corriente", "Pasivo corriente","Pasivo no corriente", "CURRENT","NON-CURRENT","CURRENT ASSETS","NON-CURRENT ASSETS","CURRENT LIABILITIES","NON-CURRENT LIABILITIES" ]
# (Para 'Utilidad' no se hace un match directo, sino que la usaremos como "por defecto")

# 2) Crear una columna "tipo_marker" que solo se llena si Descripcion es uno de los tipos
df_marcado = df_final.withColumn(
    "corriente_tipo",
    F.when(F.col("Descripcion").isin(tipos_especiales),  F.initcap(F.col("Descripcion")))
    .otherwise(None)
)

# Asegurarse de que la columna 'row_id' existe
df_marcado = df_marcado.withColumn("row_id", F.monotonically_increasing_id())

w = Window.partitionBy("Empresa").orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

df_relleno = df_marcado.withColumn(     
    "relleno_tipo",
    F.last("corriente_tipo", ignorenulls=True).over(w)
)

# 4) Si "tipo_filled" es nulo, significa que todavía no hemos visto "ACTIVO"/"PASIVO"/"PATRIMONIO"
#    => asignamos "Utilidad" como valor por defecto
df_ultimo = df_relleno.withColumn(
    "TipoCorriente",
    when(
        (
            (col("ValorContable") == 'Activo') |
            (col("ValorContable") == 'Pasivo') |
            (col("ValorContable") == 'Assets') |
            (col("ValorContable") == 'Liabilities')
        ),
         F.coalesce(F.col("relleno_tipo"))
    )\
    .otherwise("N/A")
   
)

df_ultimo = df_ultimo\
    .select(col('row_id'),col('Descripcion'),col('Empresa'),col('Valor'),col('ValorContable'), col('TipoCorriente'))

display(df_ultimo)

# COMMAND ----------

fecha = datetime.now()

#Eliminar las filas con información innecesarias

df_limpieza = df_ultimo\
    .withColumn(
        "TipoCorriente",
        F.when(
            (
                (
                    (col('ValorContable').like('Assets')) &
                    (col('TipoCorriente').like('Current'))
                ) |
                (
                    (col('ValorContable').like('Assets')) &
                    (col('TipoCorriente').like('Current Assets'))
                ) 
            ),
            "Activo Corriente"
        )\
        .otherwise(
            F.when(
                (
                    (
                        (col('ValorContable').like('Liabilities')) &
                        (col('TipoCorriente').like('Current'))
                    ) |
                    (
                        (col('ValorContable').like('Liabilities')) &
                        (col('TipoCorriente').like('Current Liabilities'))
                    ) 
                ),
                "Pasivo Corriente"
            )\
            .otherwise(
                F.when(
                    (
                        (
                            (col('ValorContable').like('Assets')) &
                            (col('TipoCorriente').like('Non-current'))
                        ) |
                        (
                            (col('ValorContable').like('Assets')) &
                            (col('TipoCorriente').like('Non-current Assets'))
                        ) 
                    ),
                    "Activo No Corriente"
                )\
                .otherwise(
                    F.when(
                        (
                            (
                                (col('ValorContable').like('Liabilities')) &
                                (col('TipoCorriente').like('Non-current'))
                            ) |
                            (
                                (col('ValorContable').like('Liabilities')) &
                                (col('TipoCorriente').like('Non-current Liabilities'))
                            ) 
                        ),
                        "Pasivo No Corriente"
                    )\
                    .otherwise(col('TipoCorriente'))
                )
            )
        )
    )\
    .filter(
        (col('Descripcion').like('%ACTIVO%') != True) &
        (col('Descripcion').like('%Activo corriente%') != True) &
        (col('Descripcion').like('%Activo no corriente%') != True) &
        (col('Descripcion').like('%PASIVO%') != True) &
        (col('Descripcion').like('%Pasivo corriente%') != True) &
        (col('Descripcion').like('%Pasivo no corriente%') != True) &
        (col('Descripcion').like('%Patrimonio%') != True) &
        (col('Descripcion').like('%Assets%') != True) &
        (col('Descripcion').like('%CURRENT%') != True) &
        (F.upper(col('Descripcion')) != F.upper(col('Empresa'))) &
        (F.upper(col('Descripcion')) != F.upper(col('ValorContable'))) &
        (~col('Descripcion').like('%(BRL thousand)%')) &
        (col('Descripcion').like('%EQUITY%') != True) 
    )\
    .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Gross", "Utilidad") )\
    .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Assets", "Activos") )\
    .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Liabilities", "Pasivos") )\
    .withColumn("ValorContable", F.regexp_replace(F.col("ValorContable"), "Equity", "Patrimonio") )\
    .filter(~col('Descripcion').isin(empresas_list))\
    .withColumn('Año', lit(full_ano))\
    .withColumn('Trimestre', lit(trimestre))\
    .withColumn("FECHA_PUBLICACION",lit(fecha))\
    .withColumn("Valor",col("Valor").cast("double"))\
    .fillna(0)\
    .withColumn(
        "Estado",
        F.trim(
                 F.split(F.col("Empresa"), "-")[0]
             )
    )\
    .withColumn(
        "Empresa",
        F.trim(
                 F.split(F.col("Empresa"), "-")[1]
        )
    )\
    .select(
        col('Estado'),
        col('Empresa'),
        col('Descripcion'),
        col('ValorContable'),
        col('TipoCorriente'),
        col('valor').cast('float'),
        col('Año').cast('int'),
        col('Trimestre').cast('int'),
        col('FECHA_PUBLICACION')
    )

display(df_limpieza)


# COMMAND ----------

# MAGIC %sql 
# MAGIC select `Año`,Trimestre, COUNT(*) AS cantidad 
# MAGIC from Delta.`/Volumes/test_data/base/deltatables/EEFF`
# MAGIC GROUP BY `Año`,Trimestre
# MAGIC

# COMMAND ----------

df_limpieza.createOrReplaceTempView("resultado")

display(
    spark.sql(
        """
            select `Año`,Trimestre,Empresa,Estado,ValorContable,TipoCorriente,Descripcion, COUNT(*) AS cantidad 
            from resultado
            GROUP BY `Año`,Trimestre,Empresa,Estado,ValorContable,TipoCorriente,Descripcion
            HAVING COUNT(*) > 1;
        """
    )
)

# COMMAND ----------

# creado = False

# if not creado:
#     df_guardado = spark.createDataFrame([], df_limpieza.schema)

# df_guardado = df_guardado.union(df_limpieza)

spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

table_name = "test_data.base.EEFF"

df_limpieza.write \
    .format("delta") \
    .mode('overwrite') \
    .partitionBy("Año","Trimestre","Empresa","Estado","ValorContable","TipoCorriente") \
    .save("/Volumes/test_data/base/deltatables/EEFF/")



if not spark.catalog.tableExists(table_name):

    spark.sql(
        """
            CREATE TABLE test_data.base.EEFF
            AS SELECT *
            FROM delta.`/Volumes/test_data/base/deltatables/EEFF/`
        """
    )
else:
    # Cargar la tabla destino desde Unity Catalog
    tabla_destino = DeltaTable.forName(spark, "test_data.base.EEFF")
    
    condicion = """
    target.`Año` = source.`Año` AND 
    target.Trimestre = source.Trimestre AND 
    target.Empresa = source.Empresa AND
    target.ValorContable = source.ValorContable AND
    target.TipoCorriente = source.TipoCorriente AND
    target.Estado = source.Estado AND
    target.Descripcion = source.Descripcion

    """

    tabla_destino.alias("target").merge(
        source = df_limpieza.alias("source"),
        condition = condicion
    ).whenMatchedUpdateAll() \
    .whenNotMatchedInsertAll() \
    .execute()

# COMMAND ----------

# %sql
# use catalog `test_data`; 

# drop table if exists `base`.`eeff`;
