# Databricks notebook source
import pyspark.sql.functions as F
from pyspark.sql.functions import col, lit,date_format, when
from pyspark.sql.window import Window
from datetime import datetime 
from openpyxl import load_workbook
import pandas as pd
from delta.tables import DeltaTable
import json

# COMMAND ----------

rutas = ['abfs:/Volumes/test_data/base/4q24/Deuda/01 Deuda_Dic2024.xlsx']
#hojas_origen = ['KitBonos']

hojas_origen = ['KitCreditos']

# COMMAND ----------

source_wb = load_workbook(rutas[0][5:], data_only=True)
source_ws = source_wb[hojas_origen[0]]

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        if str(source_ws.cell(row=fila, column=columna).value).lower() == 'filial' :
            fila_encabezado = fila
            columna_inicio = columna

df_pandas = pd.read_excel(
    rutas[0][5:],
    sheet_name=hojas_origen[0],
    engine='openpyxl',
    header=fila_encabezado - 1
    )
df_pandas = df_pandas\
        .astype(str)

display(df_pandas)

# COMMAND ----------

for v in range(len(df_pandas.columns.values)):
    if 'moneda del crédito' in df_pandas.columns.values[v]:
        fecha = df_pandas.columns.values[v].split('\n')[0]
        mes = fecha.split('-')[0]
        ano = fecha.split('-')[1]
        nombre_columna_moneda_credito = df_pandas.columns.values[v].split('\n')[1][1:-1]
        df_pandas.rename(columns={df_pandas.columns.values[v] : "MonedaDelCredito"},inplace=True) 

    if 'millones de COP' in df_pandas.columns.values[v]:
        nombre_columna_millones_cop = df_pandas.columns.values[v].split('\n')[1][1:-1]
        df_pandas.rename(columns={df_pandas.columns.values[v] : 'MillonesDeCOP'},inplace=True) 
       
    if 'miles USD' in df_pandas.columns.values[v]:
        nombre_columna_miles_USD = df_pandas.columns.values[v].split('\n')[1][1:-1]
        df_pandas.rename(columns={df_pandas.columns.values[v] : 'MilesUSD'},inplace=True) 

if hojas_origen[0] == 'KitCreditos':
    df_pandas = df_pandas.iloc[:-1,1:]

display(df_pandas)

# COMMAND ----------

df_inicial = spark.createDataFrame(df_pandas)

df_columna_filtrada = df_inicial\
    .drop(col('Unnamed: 10'), col('Unnamed: 14'), col('COP'), col('USD'))\
    .replace('nan', None)


df_base = df_columna_filtrada\
    .dropna(how="all")

display(df_base)

# COMMAND ----------

df_base = df_base\
    .filter(
        (col('Filial') != '(1) Bonos emitidos en PEN y se hizo un SWAP a USD con el Banco BBVA.') |
        (col('Filial').isNull() == True) 
    )

df_base = df_base\
    .filter(
        (col('Filial') != '(2) Bonos emitidos en USD y se hizo un SWAP a UF.') |
        (col('Filial').isNull() == True) 
    )

df_base = df_base\
    .filter(
        (col('Filial') != '(3) Opcion Call') |
        (col('Filial').isNull() == True) 
    )

df_base = df_base\
    .filter(
        (col('Filial') != 'Bonos Verdes') |
        (col('Filial').isNull() == True) 
    )

df_base = df_base\
    .filter(
        (col('Filial') != 'TOTAL BONOS') |
        (col('Filial').isNull() == True) 
    )

display(df_base)

# COMMAND ----------

df_normalizado = df_base\
    .withColumn("MesMoneda", lit(mes))\
    .withColumn("AnoMoneda", lit(ano))\
    .withColumn('FechaVencimiento', date_format(col('Fecha Vencimiento'), 'yyyy-MM-dd'))\
    .withColumn(
        "PlazoAnos",
        F.when(
            F.col("Plazo Años").isNotNull(),
            F.expr("cast(`Plazo Años` as int)")
        )
    )\
    .withColumn(
        "MonedaDelCredito",
        F.col("MonedaDelCredito").cast("decimal(20,0)")
    )\
    .withColumn(
        "MillonesDeCOP",
        F.col("MillonesDeCOP").cast("decimal(10,0)")
    )\
    .withColumn(
        "MilesUSD",
        F.col("MilesUSD").cast("decimal(10,0)")
    )\
    .withColumn("Unnamed9",F.concat_ws(" ", F.col("Unnamed: 9") * 100,lit("%")))\
    .withColumn(
        "TasaDeInteres",
        F.concat_ws(" ",F.col("Tasa de Interés"),F.col("Unnamed: 8"), F.col("Unnamed9"))
    )

display(df_normalizado)

# COMMAND ----------

if hojas_origen[0] == 'KitBonos':

    df_normalizado = df_normalizado\
        .withColumn('FechaEmision', date_format(col('Fecha Emisión'), 'yyyy-MM-dd'))
            
else:
    df_normalizado = df_normalizado\
        .withColumn('FechaEmision', date_format(col('Fecha Inicial'), 'yyyy-MM-dd'))\
        .withColumn('Tipo',
                    when(col('Fuente de Financiación').isNotNull(), lit("Creditos"))
                    )
        


df_normalizado = df_normalizado\
    .filter(
            ~((col("Filial") == " ") &
            (col("Tipo") == " "))
        )\
    .select(
            "Filial",
            "Tipo",
            col("Fuente de Financiación").alias("FuentedeFinanciacion"),
            col("Moneda Crédito").alias("MonedaCredito"),
            "FechaEmision",
            "FechaVencimiento",
            "PlazoAnos",
            "TasaDeInteres",
            "AnoMoneda",
            "MesMoneda",
            "monedadelcredito",
            "millonesdeCOP",
            "milesUSD"
        )

display(df_normalizado)

# COMMAND ----------

#Creamos un ID para preservar el orden original de las filas
df1 = df_normalizado.withColumn("row_id", F.monotonically_increasing_id())

# 2) Marcamos filas de país y filas de compañía
#    - Fila de "país": Descripcion != null, valor == null
#    - Fila de "compañía" (o su "marker"): Descripcion != null y valor != null
#      (o, si tu lógica requiere, puede ser Descripcion != null sin importar valor,
#       y luego filtras al final las que tengan valor no nulo)
df2 = df1.withColumn(
    "is_pais",
    (F.col("Filial").isNotNull()) & (F.col("Tipo").isNull())
).withColumn(
    "is_compania_marker",
    (F.col("Filial").isNotNull()) & (F.col("Tipo").isNotNull())
)

# 3) Creamos columnas 'pais_marker' y 'compania_marker' con el texto de Descripcion
#    solo en las filas correspondientes
df3 = df2.withColumn(
    "pais_marker",
    F.when(F.col("is_pais"), F.col("Filial"))
).withColumn(
    "compania_marker",
    F.when(F.col("is_compania_marker"), F.col("Filial"))
)

# 4) Definimos una ventana que recorra las filas en orden de row_id
#    y haga un "forward fill" (último valor no nulo) hasta la fila actual
w = Window.orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

df4 = df3.withColumn(
    "Pais",
    F.last("pais_marker", ignorenulls=True).over(w)
).withColumn(
    "Compania",
    F.last("compania_marker", ignorenulls=True).over(w)
)

# 5) Filtramos para quedarnos solo con las filas que tengan un valor no nulo
#    (son las filas "compañía" y las filas donde Descripcion era null pero la compañía ya está arrastrada)
df5 = df4.filter(F.col("Tipo").isNotNull())

fecha = datetime.now()
# 6) Seleccionamos las columnas finales:
df_ultimo = df5\
    .withColumn("FECHAPUBLICACION", lit(fecha))\
    .select(
    "Compania",
    "Pais",
    "Tipo",
    "FuentedeFinanciacion",
    "MonedaCredito",
    "FechaEmision",
    "FechaVencimiento",
    "PlazoAnos",
    "TasaDeInteres",
    "AnoMoneda",
    "MesMoneda",
    "monedadelcredito",
    "millonesdeCOP",
    "milesUSD",
    "FECHAPUBLICACION"
)
        

display(df_ultimo)

# COMMAND ----------

spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

table_name = "test_data.base.deuda"

df_ultimo.write \
    .format("delta") \
    .mode("overwrite") \
    .partitionBy("Pais","Compania","Tipo","FuentedeFinanciacion","AnoMoneda","MesMoneda","MonedaCredito") \
    .save("/Volumes/test_data/base/deltatables/Deuda/")

if not spark.catalog.tableExists(table_name):

    spark.sql(
        """
            CREATE TABLE test_data.base.Deuda
            AS SELECT *
            FROM delta.`/Volumes/test_data/base/deltatables/Deuda/`
        """
    )
else:
    # Cargar la tabla destino desde Unity Catalog
    tabla_destino = DeltaTable.forName(spark, "test_data.base.deuda")
    
    condicion = """
    target.Pais = source.Pais AND 
    target.Compania = source.Compania AND 
    target.Tipo = source.Tipo AND
    target.AnoMoneda = source.AnoMoneda AND 
    target.MesMoneda = source.MesMoneda AND 
    target.MonedaCredito = source.MonedaCredito AND
    target.FuentedeFinanciacion = source.FuentedeFinanciacion
    """

    tabla_destino.alias("target").merge(
        source = df_ultimo.alias("source"),
        condition = condicion
    ).whenMatchedUpdateAll() \
    .whenNotMatchedInsertAll() \
    .execute()




# COMMAND ----------

# MAGIC %sql
# MAGIC select * from delta.`/Volumes/test_data/base/deltatables/Deuda/`

# COMMAND ----------

# %sql

# SELECT `Tipo`,count(*) FROM test_data.base.deuda GROUP BY `Tipo`;

# COMMAND ----------

# %sql
# use catalog `test_data`; 
# drop table `base`.`deuda`;
