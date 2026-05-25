# Databricks notebook source
import pyspark.sql.functions as F
from pyspark.sql.functions import col, lit, when
from pyspark.sql.window import Window
from datetime import datetime 
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from pyspark.sql.types import *
from delta.tables import DeltaTable 
import json

# COMMAND ----------

rutaArchivos = ['abfs:/Volumes/test_data/base/4q24/Deuda/01 Deuda_Dic2024.xlsx', 'abfs:/Volumes/test_data/base/4q23/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx']
hojas_origen = ['PerfilD']
ano = 2024

rutaArchivo = rutaArchivos[0][5:]

# COMMAND ----------

source_wb = load_workbook(rutaArchivo, data_only=True)
source_ws = source_wb[hojas_origen[0]]
encontrado = False

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        
        valor = str(source_ws.cell(row=fila, column=columna).value).upper()
        #print(f'{fila}.{columna}) {valor}')

        if 'PERFIL DE VENCIMIENTO POR AÑOS' in valor:
            fila_encabezado = fila
            columna_inicio = columna
            encontrado = True
            break

    if encontrado:
        break

print(f'fila_encabezado: {fila_encabezado}, columna_inicio: {columna_inicio}')

# COMMAND ----------

lista_color_celda = []
#i = 1
for fila in range(fila_encabezado + 2, source_ws.max_row + 1):
    celda = source_ws.cell(row=fila, column=columna_inicio + 2)
    #print(f'{i}) {celda.value} --> {celda.fill.fgColor.tint}')
    #i = i +1
    lista_color_celda.append(celda.fill.fgColor.tint)

# COMMAND ----------

df_pandas = pd.read_excel(
    rutaArchivo,
    sheet_name= hojas_origen[0],
    engine='openpyxl',
    header=fila_encabezado - 1
    )

#df_pandas

# COMMAND ----------

col_index = df_pandas.columns.get_loc('PERFIL DE VENCIMIENTO POR AÑOS')
col_index_datos = df_pandas.columns.get_loc(ano)

df_pandas_final = df_pandas[[df_pandas.columns[col_index + 1], df_pandas.columns[col_index + 2], df_pandas.columns[col_index_datos]]].iloc[:-1,:]

#display(df_pandas_final)

# COMMAND ----------

df_inicio = spark.createDataFrame(df_pandas_final)

display(df_inicio)

# COMMAND ----------

df_colores = spark.createDataFrame(lista_color_celda,'double')

display(df_colores)

# COMMAND ----------

df_inicial = df_inicio\
    .replace('nan', None)\
    .select(
        col(df_inicio.columns[0]).alias('Compañia'),
        col(df_inicio.columns[1]).alias('Moneda'),
        col(f'{ano}').alias('Valor')
    )\
    .withColumn("Valor", F.ceil(F.col("Valor")).cast("int"))

display(df_inicial)

# COMMAND ----------

ventana = Window.orderBy(F.lit(1))
df_inicial_indexado = df_inicial.withColumn("row_id", F.row_number().over(ventana))
df_colores_indexado = df_colores.withColumn("row_id", F.row_number().over(ventana))

# Realizamos el join sobre el índice y luego eliminamos la columna row_id
df_unido = df_inicial_indexado.join(df_colores_indexado, on="row_id").drop("row_id")

df_unido = df_unido\
    .withColumn(
            "value", 
            F.when(
                    (
                        (F.col("value") == 0) &
                        (F.lag(F.col("value")).over(ventana) != 0)
                    ), 
                    F.lag(F.col("value")).over(ventana)
                ).otherwise(F.col("value"))
    )\
    .filter(
        ~(
            (col('Compañia').isNull()) &
            (col('Moneda').isNull()) &
            (col('Valor').isNull())
        )
    )

display(df_unido)

# COMMAND ----------

df_relleno = df_unido.withColumn("row_id", F.monotonically_increasing_id())

ventana_contexto_relleno = (
    Window
    .orhttps://adb-1074487248609599.19.azuredatabricks.net/editor/notebooks/1547090052092464?o=1074487248609599$0derBy("row_id")
    .rowsBetween(1, Window.unboundedFollowing)  # de la fila siguiente hasta la última
)

df_relleno = df_relleno\
    .withColumn(
    "Pais",
    F.when(
        F.col("value") != 0,            # Si la fila actual no es cero
        F.col("Compañia")                  # usamos el propio valor
    ).otherwise(
        # Si la fila actual es cero, buscamos en las filas siguientes:
        F.first(                        # tomamos el primer Valor no-cero
            F.when(F.col("Value") != 0, F.col("Compañia")),  # si es 0 => null
            ignorenulls=True
        ).over(ventana_contexto_relleno)
    )
)\
.withColumn(
    "Pais",
    when(
        col("Pais").isNull(),
        F.lag("Pais").over(ventana)
    ).otherwise(col("Pais"))
)\
    .filter(
        (col('Pais').isNotNull()) &
        (col('Pais') != 'TOTAL DEUDA')
    )
    


#display(df_relleno)

# COMMAND ----------

ventana_valor_anterior = Window.orderBy("row_id")
fecha = datetime.now()

df_tratado = df_relleno\
    .filter(
        (col('Pais').isNotNull()) &
        (col('Pais') != 'TOTAL DEUDA')
    )\
    .withColumn(
        "Compañia",
        F.when(
            F.col("Compañia").isNotNull(),            # Si la fila actual no es cero
            F.col("Compañia")                  # usamos el propio valor
        ).otherwise(
             F.lag("Compañia", 1).over(ventana_valor_anterior)
            
        )
    )\
    .withColumn(
        'Año',
        lit(f'{ano}')
    )\
    .withColumn(
        'FECHAPUBLICACION',
        lit(fecha.strftime('%Y-%m-%d'))
    )\
    .filter(
        col('Pais') != col('Compañia')
    )

df_final = df_tratado\
    .select(
        col('Pais'),
        col('Compañia').alias('Compania'),
        col('Moneda'),
        col('Valor'),
        col('Año').alias('Ano').cast('int'),
        col('FECHAPUBLICACION'),
    )
    

display(df_final)

# COMMAND ----------

spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

table_name = "test_data.base.PerfilDeuda"

df_final.write \
    .format("delta") \
    .mode("overwrite") \
    .partitionBy("Ano","Moneda","Pais","Compania") \
    .save("/Volumes/test_data/base/deltatables/PerfilDeuda/")

if not spark.catalog.tableExists(table_name):

    spark.sql(
        """
            CREATE TABLE test_data.base.PerfilDeuda
            AS SELECT *
            FROM delta.`/Volumes/test_data/base/deltatables/PerfilDeuda/`
        """
    )
else:
    # Cargar la tabla destino desde Unity Catalog
    tabla_destino = DeltaTable.forName(spark, "test_data.base.PerfilDeuda")
    
    condicion = """
    target.Pais = source.Pais AND 
    target.Compania = source.Compania AND 
    target.Moneda = source.Moneda AND
    target.Ano = source.Ano 
    """

    tabla_destino.alias("target").merge(
        source = df_final.alias("source"),
        condition = condicion
    ).whenMatchedUpdateAll() \
    .whenNotMatchedInsertAll() \
    .execute()

# COMMAND ----------

# %sql
# SELECT * FROM DELTA.`/Volumes/test_data/base/deltatables/PerfilDeuda/`;

# COMMAND ----------

# %sql
# drop table test_data.base.perfildeuda
