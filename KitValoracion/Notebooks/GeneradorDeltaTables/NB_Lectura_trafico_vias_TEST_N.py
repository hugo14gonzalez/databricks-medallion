# Databricks notebook source
import pandas as pd
from pyspark.sql.functions import col, lit
from datetime import datetime 
from delta.tables import DeltaTable
from openpyxl import load_workbook
import re
import json

# COMMAND ----------

rutaArchivos = ['/Volumes/brz_lz_crudos/isa_transversal/vol_estructured_files/KitValoracion/2024/3Q24/Trafico vías/InfoVías-ReportesTrimestralesISA-paraRI.xlsx']
hojas_origen = ['Tráfico vías']
full_ano = 2024
trimestre = 3
ano = int(str(full_ano)[2:])

# COMMAND ----------

source_wb = load_workbook(rutaArchivos[0], data_only=True)

# Obtiene la lista de nombres de hoja
nombres_hojas = source_wb.sheetnames

# Muestra los nombres
print(nombres_hojas)

source_ws = source_wb[hojas_origen[0]]


# COMMAND ----------

#source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
#source_ws = source_wb[hojas_origen[0]]
encontrado = False

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        
        valor = str(source_ws.cell(row=fila, column=columna).value).replace(' ','')
        #print(f'{fila}.{columna}) {valor}')

        if str(full_ano) == valor\
        or bool(re.match(r"^\d{4}-[1-4][T,Q]$",valor))\
        or bool(re.match(r"^[1-4][T,Q]\d{2}$",valor)):
            fila_encabezado = fila - 1
            columna_inicio = columna - 1
            encontrado = True
            break

    if encontrado:
        break

print(f'fila_encabezado: {fila_encabezado}, columna_inicio: {columna_inicio}')

# COMMAND ----------

df_pandas = pd.read_excel(
    rutaArchivos[0],
    sheet_name= hojas_origen[0],
    engine='openpyxl',
    header=fila_encabezado
    )

df_pandas = df_pandas\
            .astype(str)\
            .iloc[fila_encabezado-3:, 1:]


display(spark.createDataFrame(df_pandas))

# COMMAND ----------

df = spark.createDataFrame(df_pandas)

display(df)

# COMMAND ----------

existe = True

if f'{trimestre}Q{ano}' in df.columns: 
    valor = f'{trimestre}Q{ano}'
elif f'{full_ano} - {trimestre}Q' in df.columns:
    valor = f'{full_ano} - {trimestre}Q'
elif full_ano in df.columns:
    valor = full_ano
else:
    print(f"No existe la columna")
    existe = False

print(valor)


# COMMAND ----------

if existe:
    df_inicial = df\
        .withColumnRenamed('Unnamed: 1', 'Empresa')\
        .filter(
                (~col('Empresa').like('*TMDE: Tráfico medio diario equivalente'))
            )\
        .filter(
            (~col('Empresa').like('TMDE 2015- 2021 con base 365 días, excepto 2016 y 2020 los cuales están con base 366 días')) 
        )\
        .filter(
            (~col('Empresa').like('**TPD: Tráfico Promedio Diario') )
        )\
        .replace('nan', None)\
        .withColumn('Valor', col(f'{valor}').cast('float'))\
        .na.fill(0, ["Valor"])\
        .select(col('Empresa'), col('Valor'))
        
    display(df_inicial)

# COMMAND ----------

from pyspark.sql import SparkSession, functions as F
from pyspark.sql.window import Window

spark = SparkSession.builder.getOrCreate()

# Supongamos que tu DataFrame original se llama df y tiene columnas:
# ["Empresa", "Valor"]

# 1) Crear un ID para preservar el orden original
df1 = df_inicial.withColumn("row_id", F.monotonically_increasing_id())

# 2) Definir si la fila es "empresa" o "tipo"
#    Usamos una expresión regular que detecta "RUTA", "TMDE" o "TPD" (ignora mayúsculas).
#    Ajusta si tienes más patrones.
df2 = df1.withColumn(
    "is_empresa",
    F.col("Empresa").rlike("(?i)(RUTA|TMDE|TPD)")
).withColumn(
    "is_tipo",
    ~F.col("is_empresa")
)

# 3) "enterprise_marker": en las filas de empresa, guardamos el texto. En otras, null.
df3 = df2.withColumn(
    "enterprise_marker",
    F.when(F.col("is_empresa"), F.col("Empresa"))
)

# 4) Forward fill de la empresa: para cada fila, tomar el último valor no nulo de "enterprise_marker"
w = Window.orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

df4 = df3.withColumn(
    "empresa_filled",
    F.last("enterprise_marker", ignorenulls=True).over(w)
)

# 5) Definir las columnas finales
#    - Empresa = la empresa propagada
#    - Tipo    = el texto de la fila SÓLO si es tipo
#    - Valor = "Valor" SÓLO si es tipo
df5 = df4.withColumn(
    "Compañia",
    F.col("empresa_filled")
).withColumn(
    "Tipo",
    F.when(F.col("is_tipo"), F.col("Empresa"))
).withColumn(
    "Valor",
    F.when(F.col("is_tipo"), F.col("Valor"))
)

# 6) Filtrar para quedarse solo con las filas de tipo (descartamos las filas marcadas como empresa)
df_result = df5.filter(F.col("is_tipo"))

# 7) Seleccionar columnas finales y mostrar
df_transformado = df_result.select(col("Compañia").alias('Compania'), "Tipo", "Valor")



display(df_transformado)


# COMMAND ----------

fecha = datetime.now()

df_final = df_transformado\
    .withColumn("Año", lit(f'{full_ano}'))\
    .withColumn("Trimestre", lit(f'{trimestre}'))\
    .withColumn("FECHAPUBLICACION", lit(f'{fecha}'))\
    .na.fill(0, ["valor"])

df_final = df_final.select(
    col('Compania').alias('Compania'),
    col('Tipo'),
    col('Valor'),
    col('Año').alias('Ano'),
    col('Trimestre'),
    col('FECHAPUBLICACION')
)

display(df_final)


# COMMAND ----------

spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

table_name = "test_data.base.Trafico_vias"

df_final.write \
    .format("delta") \
    .mode("overwrite") \
    .partitionBy("Ano","Trimestre","Compania","Tipo") \
    .save("/Volumes/test_data/base/deltatables/Trafico_vias/")


if not spark.catalog.tableExists(table_name):

    spark.sql(
        """
            CREATE TABLE test_data.base.Trafico_vias
            AS SELECT *
            FROM delta.`/Volumes/test_data/base/deltatables/Trafico_vias/`
        """
    )
else:
    # Cargar la tabla destino desde Unity Catalog
    tabla_destino = DeltaTable.forName(spark, "test_data.base.Trafico_vias")
    
    condicion = """
    target.Compania = source.Compania AND 
    target.Ano = source.Ano AND
    target.Tipo = source.Tipo AND
    target.Trimestre = source.Trimestre
    """

    tabla_destino.alias("target").merge(
        source = df_final.alias("source"),
        condition = condicion
    ).whenMatchedUpdateAll() \
    .whenNotMatchedInsertAll() \
    .execute()

# COMMAND ----------

# %sql
# drop table test_data.base.Trafico_vias

