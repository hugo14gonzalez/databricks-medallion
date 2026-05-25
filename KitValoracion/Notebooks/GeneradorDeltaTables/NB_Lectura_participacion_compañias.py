# Databricks notebook source
import pandas as pd
import pyspark.sql.functions as F
import json
import unicodedata
from openpyxl import load_workbook
from pyspark.sql.functions import col, lit, when, regexp_replace, lower,udf
from pyspark.sql.window import Window
from datetime import datetime
from delta.tables import DeltaTable
from pyspark.sql.types import StringType

# COMMAND ----------

# ruta = ['abfs:/Volumes/test_data/base/4q23/Participaciones Empresas/RP_Participacion Directa & Efectiva0924_3Q_2024.xlsm']
# hojas_origen = ['Reporte_SP']

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
catalogo = dbutils.widgets.get('CatalogoDelta')
esquema = dbutils.widgets.get('EsquemaDelta')

#ruta = f'/Volumes/{catalogo}/{esquema}/'
table_name = f"{catalogo}.{esquema}.tbl_participacion_empresas"

df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn('Carpeta', normalizar_udf(col('Carpeta')) )\
    .filter(
        (col('Carpeta') == 'participaciones empresas') 
    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]

# COMMAND ----------

try:
    # Cargar los libros de Excel
    source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
    source_ws = source_wb[hojas_origen[0]]
    print(f'Filas del archivo: {source_ws.max_row}')

    for fila in range(1, source_ws.max_row + 1):
        for columna in range(1, 5):
            valor = str(source_ws.cell(row=fila, column=columna).value).lower()
            #print(f'{fila}.{columna}) {valor}')
            if valor == 'tipo de \ncontrol' :
                fila_encabezado = fila - 1
                columna_inicio = columna - 1
    
    print(f'Fila encabezado: {fila_encabezado}, Columna inicio: {columna_inicio}')

    df_pandas = pd.read_excel(
        rutaArchivos[0][5:],
        sheet_name=hojas_origen[0],
        engine='openpyxl',
        header=fila_encabezado
    )
    df_pandas = df_pandas.iloc[:,columna_inicio:]
    df_pandas = df_pandas.astype(str)
            
    df = spark.createDataFrame(df_pandas)
    df = df\
        .select(
            col('Tipo de \nControl').alias('TipoDeControl'), col('Negocio'), col('País'), col('SociedadGL'), 
            col('% Directo ISA').alias('PorcentajeDirectoISA'), col('% Indirecto a través de Otras Filiales').alias('PorcentajeIndirectoOtrasFiliales'), 
            col('% Efectivo ISA').alias('PorcentajeEfectivoISA'), col('SocIndirecta'))\
        .replace('nan', None)

    display(df)

    # Agregar un índice para mantener el orden (puedes ajustar este método según tu fuente de datos)
    df = df.withColumn("id_fila", F.monotonically_increasing_id())

    # Definir una ventana ordenada por 'id'
    ventana_orden = Window.orderBy("id_fila")

    # Definir una ventana que ordene por el índice y que abarque desde la primera fila hasta la actual
    ventana_entre_datos = Window.orderBy("id_fila").rowsBetween(Window.unboundedPreceding, 0)

    df_ultimo = df\
        .withColumn("SociedadGL", F.last("SociedadGL", ignorenulls=True).over(ventana_entre_datos))\
        .withColumn("País", F.last("País", ignorenulls=True).over(ventana_entre_datos))\
        .withColumn("Negocio", F.last("Negocio", ignorenulls=True).over(ventana_entre_datos))\
        .withColumn("TipoDeControl", F.last("TipoDeControl", ignorenulls=True).over(ventana_entre_datos))\
        .withColumn(
            "PorcentajeDirectoISA", 
            when(
                (
                    col("SociedadGL").isNull() &
                    F.lag("PorcentajeDirectoISA", 1).over(ventana_orden).isNotNull() 
                ),
                F.last("PorcentajeDirectoISA", ignorenulls=True).over(ventana_entre_datos)
            )\
            .otherwise(col('PorcentajeDirectoISA'))
            
        )

    windowSpec = (
        Window.partitionBy("SociedadGL")
            .orderBy("id_fila")
            .rowsBetween(Window.unboundedPreceding, Window.unboundedFollowing)
    )
    columns_to_fill = [c for c in df_ultimo.columns if c not in ("SociedadGL", "id_fila")]

    for c in columns_to_fill:
        # 4a. Creamos una columna temporal con el valor de la primera fila del grupo
        first_col = c + "_first"
        df_ultimo = df_ultimo.withColumn(first_col, F.first(c, ignorenulls=False).over(windowSpec))
        
        # 4b. Rellenamos la columna original con ese valor si está en null
        df_ultimo = df_ultimo.withColumn(c, F.coalesce(F.col(c), F.col(first_col)))
        

    df_limpio = df_ultimo\
        .select(
        col('TipoDeControl'),
        col('Negocio'),
        col('País').alias('Pais'),
        col('SociedadGL'),
        col('PorcentajeDirectoISA'),
        col('PorcentajeIndirectoOtrasFiliales'),
        col('PorcentajeEfectivoISA'),
        col('SocIndirecta')
    )

    display(df_limpio)
    
    fecha = datetime.now()
    
    df_final = df_limpio \
        .withColumn("PorcentajeDirectoISA", F.when(F.col("PorcentajeDirectoISA").isNotNull(), F.concat_ws(" ", F.round(F.col("PorcentajeDirectoISA") * 100.0, 6), lit("%")))) \
        .withColumn("PorcentajeIndirectoOtrasFiliales", F.when(F.col("PorcentajeIndirectoOtrasFiliales").isNotNull(), F.concat_ws(" ", F.round(F.col("PorcentajeIndirectoOtrasFiliales") * 100.0, 6), lit("%")))) \
        .withColumn("PorcentajeEfectivoISA",  F.when(F.col("PorcentajeEfectivoISA").isNotNull(), F.concat_ws(" ", F.round(F.col("PorcentajeEfectivoISA") * 100.0, 6), lit("%")))) \
        .withColumn("AnoConsulta", lit(full_ano)) \
        .withColumn("TrimestreConsulta", lit(trimestre)) \
        .withColumn("SociedadGL", regexp_replace(col("SociedadGL"), r"\.$", ""))\
        .withColumn("SocIndirecta", regexp_replace(col("SocIndirecta"), r"\.$", ""))\
        .withColumn("FECHAPUBLICACION", lit(fecha))

    display(df_final)

    # Not supported in serverless, comment out
    #spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

    if not spark.catalog.tableExists(table_name):

        df_final.write \
            .format("delta") \
            .mode('overwrite') \
            .partitionBy("AnoConsulta","TrimestreConsulta","TipoDeControl","Negocio","Pais","SociedadGL","SocIndirecta") \
            .saveAsTable(table_name)
    else:
        df_final.createOrReplaceTempView("tmp_participacion_empresas")
        spark.sql(
            f"""
                MERGE INTO {table_name} AS target
                    USING tmp_participacion_empresas AS source
                    ON  target.AnoConsulta       = source.AnoConsulta
                    AND target.TrimestreConsulta   = source.TrimestreConsulta
                    AND target.TipoDeControl       = source.TipoDeControl
                    AND target.Negocio    = source.Negocio
                    AND target.Pais    = source.Pais
                    AND target.SociedadGL    = source.SociedadGL
                    AND target.SocIndirecta    = source.SocIndirecta

                    WHEN MATCHED THEN
                    UPDATE SET *
                    WHEN NOT MATCHED THEN
                    INSERT *;
            """
        )
except Exception as ex:
    print('Task Fail--> ### ' + str(ex) + ' ###')
