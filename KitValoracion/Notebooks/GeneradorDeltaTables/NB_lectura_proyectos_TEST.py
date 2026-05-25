# Databricks notebook source
#validaciones
    #año_reporte == full_ano
    #mes_reporte == nombre_mes
    #tipo 1 != Renovación
    #Nombre del proyecto != Proyectos de refuerzos en ejecución

#transformacion
    #fecha energizacion RI --> sacar trimestre y año
    #tomar columnas:
        # Empresa
        # Pais
        # Tipo control
        # Fecha energización RI
        # Trimestre
        # Año
        # Nombre del proyecto
        # Inversión aprobada[USD] millones USD
        # Comentarios
    #REP transformar a CTM

# COMMAND ----------

import pandas as pd
from datetime import datetime 
from pyspark.sql.functions import col, col, year, quarter,expr, lit
from openpyxl import load_workbook
from pyspark.sql.types import StructType, StructField, DateType, TimestampType
import pyspark.sql.functions as F
from delta.tables import DeltaTable
import json

# COMMAND ----------

rutaArchivos = ['abfs:/Volumes/test_data/base/4q23/Proyectos/Información Negocio TE_3Q_2024.xlsx']
hojas_origen = ['Construcción proyetos']
full_ano = 2024
trimestre = '4'

# COMMAND ----------

# lectura_carpetas = dbutils.widgets.get('df_base')
# df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
# df_carpetas = df_carpetas\
#     .withColumnRenamed('_1','Carpeta')\
#     .withColumnRenamed('_2','RutaArchivo')\
#     .withColumnRenamed('_3','HojaOrigen')\
#     .withColumnRenamed('_4','HojaDestino')\
#     .filter(
#         (col('Carpeta') == 'Proyectos')
#     )
    
# ano = dbutils.widgets.get('Anio')
# trimestre = dbutils.widgets.get('Trimestre')
# rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
# hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]

# COMMAND ----------

if trimestre == '1':
    mes = 'Marzo'
elif trimestre == '2':
    mes = 'Junio'
elif trimestre == '3':
    mes = 'Septiembre'
else:
    mes = 'Diciembre'

print(mes)

# COMMAND ----------

source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
source_ws = source_wb[hojas_origen[0]]
encontrado = False

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        
        valor = str(source_ws.cell(row=fila, column=columna).value).lower().replace('\n','')
        #print(f'{fila}.{columna}) {valor}')

        if 'añoreporte' in valor:
            fila_encabezado = fila - 1
            columna_inicio = columna - 1
            encontrado = True
            break

    if encontrado:
        break

print(f'fila_encabezado: {fila_encabezado}, columna_inicio: {columna_inicio}')

# COMMAND ----------

df_pandas = pd.read_excel(
    rutaArchivos[0][5:],
    sheet_name= hojas_origen[0],
    engine='openpyxl',
    header=fila_encabezado
    )
df_pandas = df_pandas\
            .astype(str)

display(df_pandas)

# COMMAND ----------

df_inicial = spark.createDataFrame(df_pandas)

display(df_inicial)

# COMMAND ----------

fecha = datetime.now()

df_final = df_inicial\
    .withColumnRenamed('Año\nreporte', 'Año reporte')\
    .withColumnRenamed('Inversión aprobada\n[USD]', 'Inversión aprobada USD')\
    .filter(
        (col('Año reporte') == full_ano) &
        (col('Mes reporte') == mes) &
        (col('Tipo 1') != 'Renovación') &
        (col('Nombre del proyecto') != 'Proyectos de refuerzos en ejecución')
    )\
    .replace('nan',None)\
    .withColumn("Fecha Energización RI", F.col("Fecha Energización RI").cast("date"))\
    .withColumn("Inversión aprobada USD", F.col("Inversión aprobada USD").cast("float"))\
    .withColumn("FECHAPUBLICACION", lit(fecha))\
    .withColumn("Año", year(col("Fecha Energización RI"))) \
    .withColumn("Trimestre", quarter(col("Fecha Energización RI")))\
    .select(
        col('País').alias('Pais'),
        col('Empresa').alias('Compania'),
        col('Año').alias('Ano'),
        col('Trimestre'),
        col('Tipo control').alias('TipoControl'),
        col('Inversión aprobada USD').alias('InversionAprobadaUSD'),
        col('Comentarios'),
        col('FECHAPUBLICACION')
    )

display(df_final)
    

# COMMAND ----------

df_final = df_final.groupBy(
    "Pais", "Compania", "Ano", "Trimestre", "TipoControl"
).agg(
    F.sum("InversionAprobadaUSD").alias("InversionAprobadaUSD"),
    F.first("Comentarios").alias("Comentarios")
).withColumn(
    "FECHAPUBLICACION", F.lit(fecha)
)


# COMMAND ----------

spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

table_name = "test_data.base.Proyectos"

df_final.write \
    .format("delta") \
    .mode("overwrite") \
    .partitionBy("Pais","Compania","Ano", "Trimestre", "TipoControl") \
    .save("/Volumes/test_data/base/deltatables/Proyectos/")

if not spark.catalog.tableExists(table_name):

    spark.sql(
        """
            CREATE TABLE test_data.base.Proyectos
            AS SELECT *
            FROM delta.`/Volumes/test_data/base/deltatables/Proyectos/`
        """
    )
else:
    # Cargar la tabla destino desde Unity Catalog
    tabla_destino = DeltaTable.forName(spark, "test_data.base.Proyectos")
    
    condicion = """
    target.Pais = source.Pais AND 
    target.Compañia = source.Compañia AND 
    target.Ano = source.Ano AND
    target.Trimestre = source.Trimestre AND
    target.TipoControl = source.TipoControl
    """

    tabla_destino.alias("target").merge(
        source = df_final.alias("source"),
        condition = condicion
    ).whenMatchedUpdateAll() \
    .whenNotMatchedInsertAll() \
    .execute()

# COMMAND ----------

# %sql
# drop table test_data.base.Proyectos
