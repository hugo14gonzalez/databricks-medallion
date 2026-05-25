# Databricks notebook source
import pandas as pd
import unicodedata
import json
from pyspark.sql.window import Window
from pyspark.sql.functions import col, when, count, last, lit, regexp_replace, lower,udf
from datetime import datetime 
from pyspark.sql import functions as F
from openpyxl import load_workbook
from delta.tables import DeltaTable
from pyspark.sql.types import StringType

# COMMAND ----------

# ruta = '/Volumes/brz_lz_crudos/isa_transversal/vol_estructured_files/KitValoracion/2024/3Q24/CAPEX/Capex_4Q23_Kit.xlsx'
# hoja = 'CapexPorEmpresa'

# COMMAND ----------

def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn(
        'Carpeta',
       normalizar_udf(col('Carpeta'))
    )\
    .filter(
        (col('Carpeta') == 'capex') 
    )

rutaArchivo = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]

ruta = rutaArchivo[0][5:]
hoja = hojas_origen[0]

full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')

# COMMAND ----------

source_wb = load_workbook(ruta, data_only=True)
source_ws = source_wb[hojas_origen[0]]

# COMMAND ----------

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        if str(source_ws.cell(row=fila, column=columna).value).lower() == 'capex consolidado' :
            fila_encabezado = fila
            columna_inicio = columna

# COMMAND ----------

df_pandas = pd.read_excel(
    ruta,
    sheet_name=hoja,
    engine='openpyxl',
    header=fila_encabezado - 1
    )
#display(df_pandas)

# COMMAND ----------

df = spark.createDataFrame(df_pandas)


# COMMAND ----------

existe = True
if full_ano in df.columns or str(full_ano) in df.columns:
    valor = full_ano
else:
    print(f"No existe la columna")
    existe = False

# COMMAND ----------

if existe:
    df_resultado = df\
        .select(df.columns[:-1])\
        .select(col('CAPEX CONSOLIDADO'),col(f'{valor}').alias('Valor'))


# COMMAND ----------

# 1) Agregar columna de índice
df_id = df_resultado.withColumn("id_fila", F.monotonically_increasing_id())

# 2) Crear una ventana que ordene por ese índice
windowSpec = Window.orderBy("id_fila")

# 3) Agregar columna con el valor de 'colX' de la siguiente fila
df_siguiente = df_id.withColumn("siguiente_columna", F.lead("CAPEX CONSOLIDADO").over(windowSpec))

# 4) Filtrar filas donde la siguiente fila (next_colX) NO sea nula
df_filtrado = df_siguiente.filter(F.col("siguiente_columna").isNotNull())

# 5) Eliminar la columna auxiliar si no la necesitas
df_filtrado = df_filtrado.drop("siguiente_columna", "id_fila")



# COMMAND ----------

df_tratado = df_filtrado\
    .withColumnRenamed('CAPEX CONSOLIDADO','EMPRESA')\
    .filter(col('EMPRESA').isNotNull())



# COMMAND ----------

# Crear un identificador de grupo para bloques separados por null
ventana_empresa = Window.orderBy(lit(1)).rowsBetween(Window.unboundedPreceding, 0)

# Crear una columna para determinar el bloque actual
df_tratado = df_tratado.withColumn("grupo", count(when(col("Valor").isNull(), 1)).over(ventana_empresa))



# COMMAND ----------

display(
    df_tratado\
    .withColumn(
        "PAIS",
        when(col('Valor').isNull(), col("EMPRESA"))
        .otherwise(None)
    )\
    .withColumn("PAIS", last("PAIS", True).over(ventana_empresa))\
    .filter(
        (col('Valor').isNotNull()) 
    )
)

# COMMAND ----------

df_ultimo = df_tratado\
    .withColumn(
        "PAIS",
        when(col('Valor').isNull(), col("EMPRESA"))
        .otherwise(None)
    )\
    .withColumn("PAIS", last("PAIS", True).over(ventana_empresa))\
    .filter(
        (col('Valor').isNotNull()) 
    )\
    .filter(
        ~(col('EMPRESA').like('%Inversiones dentro balance%')) 
    )\
    .filter(
        ~(col('EMPRESA').like('%Negocios que no consolidan%')) 
    )\
    .withColumn('AnoEvaluado', lit(f'{full_ano}'))\
    .withColumn("trimestreEvaluado",lit(f'{trimestre}'))\
    .withColumn("FECHA_PUBLICACION",lit(datetime.now()))\
    .select(col('PAIS'),col('EMPRESA'),col('Valor'), col('AnoEvaluado'),col("trimestreEvaluado"),col('FECHA_PUBLICACION'))


# COMMAND ----------

#display(df_ultimo)

# COMMAND ----------

# Not supported in serverless, comment out
#spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

catalogo = dbutils.widgets.get('CatalogoDelta')
esquema = dbutils.widgets.get('EsquemaDelta')
table_name = f"{catalogo}.{esquema}.tbl_capex"


if not spark.catalog.tableExists(table_name):

    df_ultimo.write \
        .format("delta") \
        .mode('overwrite') \
        .partitionBy("AnoEvaluado","trimestreEvaluado","PAIS","EMPRESA") \
        .saveAsTable(table_name)

else:

    df_ultimo.createOrReplaceTempView("tmp_capex")
    spark.sql(
       f"""
           MERGE INTO {table_name} AS target
            USING tmp_capex AS source
            ON  target.AnoEvaluado       = source.AnoEvaluado
            AND target.trimestreEvaluado = source.trimestreEvaluado
            AND target.PAIS              = source.PAIS
            AND target.EMPRESA           = source.EMPRESA

            WHEN MATCHED THEN
            UPDATE SET *
            WHEN NOT MATCHED THEN
            INSERT *;
       """
   )



# if not spark.catalog.tableExists(table_name):

#     df_ultimo.write \
#         .format("delta") \
#         .mode('overwrite') \
#         .partitionBy("AnoEvaluado","trimestreEvaluado","PAIS","EMPRESA") \
#         .save(table_name)

# #    spark.sql(
# #        f"""
# #            CREATE TABLE {table_name}
# #            AS SELECT *
# #            FROM delta.`{ruta}`
# #        """
# #    )

# else:
#     # Cargar la tabla destino desde Unity Catalog
#     #tabla_destino = DeltaTable.forName(spark, f"{table_name}")
#     tabla_destino = spark.table(table_name)
    
#     condicion = """
#     target.AnoEvaluado = source.AnoEvaluado AND 
#     target.trimestreEvaluado = source.trimestreEvaluado AND 
#     target.PAIS = source.PAIS AND 
#     target.EMPRESA = source.EMPRESA
#     """

#     tabla_destino.alias("target").merge(
#         source = df_ultimo.alias("source"),
#         condition = condicion
#     ).whenMatchedUpdateAll() \
#     .whenNotMatchedInsertAll() \
#     .execute()

# COMMAND ----------

# %sql

# SELECT `Año`,count(*) FROM test_data.base.capex GROUP BY `Año`;


# COMMAND ----------

# %sql

# SELECT * FROM test_data.base.capex ;

# COMMAND ----------

# %sql
# use catalog `test_data`; drop table `base`.`capex`;
