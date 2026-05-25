# Databricks notebook source
#Hugo
# %pip install --quiet openpyxl
# dbutils.library.restartPython()

# COMMAND ----------

import pandas as pd
import re
import json
import unicodedata
from pyspark.sql.functions import col, lit, regexp_replace, lower, udf, when, trim
from datetime import datetime 
from delta.tables import DeltaTable
from openpyxl import load_workbook
from pyspark.sql.types import StringType

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
catalogo = dbutils.widgets.get('CatalogoDelta')
esquema = dbutils.widgets.get('EsquemaDelta')

#ruta = f'/Volumes/{catalogo}/{esquema}/'
table_name = f"{catalogo}.{esquema}.tbl_trafico_vias"
ano = int(str(full_ano)[2:])

df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn('Carpeta', normalizar_udf(col('Carpeta')))\
    .filter(
        (col('Carpeta') == 'trafico vias')

    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]

# COMMAND ----------

source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
source_ws = source_wb[hojas_origen[0]]
encontrado = False

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        
        valor = str(source_ws.cell(row=fila, column=columna).value).replace(' ','')
        #print(f'{fila}.{columna}) {valor}')

        if str(full_ano) == valor\
        or bool(re.match(r"^\d{4}-[1-4][T,Q]$",valor))\
        or bool(re.match(r"^[1-4][T,Q]\d{2}$",valor)):
            fila_encabezado = fila - 1
            columna_inicio = columna - 1
            encontrado = True
            break

    if encontrado:
        break

print(f'fila_encabezado: {fila_encabezado}, columna_inicio: {columna_inicio}')

df_pandas = pd.read_excel(
    rutaArchivos[0][5:],
    sheet_name= hojas_origen[0],
    engine='openpyxl',
    header=fila_encabezado
)

df_pandas = df_pandas.astype(str).iloc[fila_encabezado-3:, 1:]
df = spark.createDataFrame(df_pandas)

# COMMAND ----------

existe = True

if f'{trimestre}Q{ano}' in df.columns: 
    valor = f'{trimestre}Q{ano}'
elif f'{full_ano} - {trimestre}Q' in df.columns:
    valor = f'{full_ano} - {trimestre}Q'
elif full_ano in df.columns:
    valor = full_ano
else:
    print(f"No existe la columna")
    existe = False

if existe:
    df_inicial = df\
        .withColumnRenamed('Unnamed: 1', 'Empresa')\
        .filter(
                (~col('Empresa').like('*TMDE: Tráfico medio diario equivalente'))
            )\
        .filter(
            (~col('Empresa').like('TMDE 2015- 2021 con base 365 días, excepto 2016 y 2020 los cuales están con base 366 días')) 
        )\
        .filter(
            (~col('Empresa').like('**TPD: Tráfico Promedio Diario') )
        )\
        .replace('nan', None)\
        .withColumn('Empresa', regexp_replace(col('Empresa'), "[-\\s\\t\\n\\r\\x0b\\x0c\\u00A0]+", ""))\
        .withColumn('Valor2', regexp_replace(col(f'{valor}'), "[-\\s\\t\\n\\r\\x0b\\x0c\\u00A0]+", ""))\
        .withColumn('Valor', (when((trim(col('Valor2')) == ""), None).otherwise(col('Valor2')))  )\
        .withColumn('Valor', (when(( (~col('Empresa').isNull()) & (trim(col('Empresa')) != "")) & 
                                   (col('Valor').isNull()), '0').otherwise(col('Valor')))  )\
        .select(col('Empresa'), col('Valor'))\
        .dropna()\
        .na.fill(0, ["Valor"])

# COMMAND ----------

from pyspark.sql import SparkSession, functions as F
from pyspark.sql.window import Window

spark = SparkSession.builder.getOrCreate()

# Supongamos que tu DataFrame original se llama df y tiene columnas:
# ["Empresa", "Valor"]

# 1) Crear un ID para preservar el orden original
df1 = df_inicial.withColumn("row_id", F.monotonically_increasing_id())

# 2) Definir si la fila es "empresa" o "tipo"
#    Usamos una expresión regular que detecta "RUTA", "TMDE" o "TPD" (ignora mayúsculas).
#    Ajusta si tienes más patrones.
df2 = df1.withColumn(
    "is_empresa",
    F.col("Empresa").rlike("(?i)(RUTA|TMDE|TPD)")
).withColumn(
    "is_tipo",
    ~F.col("is_empresa")
)

# 3) "enterprise_marker": en las filas de empresa, guardamos el texto. En otras, null.
df3 = df2.withColumn("enterprise_marker", F.when(F.col("is_empresa"), F.col("Empresa")))

# 4) Forward fill de la empresa: para cada fila, tomar el último valor no nulo de "enterprise_marker"
w = Window.orderBy("row_id").rowsBetween(Window.unboundedPreceding, 0)

df4 = df3.withColumn("empresa_filled",  F.last("enterprise_marker", ignorenulls=True).over(w))

# 5) Definir las columnas finales
#    - Empresa = la empresa propagada
#    - Tipo    = el texto de la fila SÓLO si es tipo
#    - Valor = "Valor" SÓLO si es tipo
df5 = df4.withColumn("Compañia", F.col("empresa_filled"))\
.withColumn("Tipo", F.when(F.col("is_tipo"), F.col("Empresa")))\
.withColumn("Valor", F.when(F.col("is_tipo"), F.col("Valor")))

# 6) Filtrar para quedarse solo con las filas de tipo (descartamos las filas marcadas como empresa)
df_result = df5.filter(F.col("is_tipo"))

# 7) Seleccionar columnas finales y mostrar
df_transformado = df_result.select(col("Compañia").alias('Compania'), "Tipo", "Valor")

# COMMAND ----------

try:
    fecha = datetime.now()

    df_final = df_transformado\
        .withColumn("Año", lit(f'{full_ano}'))\
        .withColumn("Trimestre", lit(f'{trimestre}'))\
        .withColumn("FECHAPUBLICACION", lit(f'{fecha}'))\
        .na.fill(0, ["valor"])

    df_final = df_final.select(
        col('Compania').alias('Compania'),
        col('Tipo'),
        col('Valor'),
        col('Año').alias('Ano'),
        col('Trimestre'),
        col('FECHAPUBLICACION')
    )

    # Not supported in serverless, comment out
    #spark.conf.set("spark.sql.sources.partitionOverwriteMode", "dynamic")

    if not spark.catalog.tableExists(table_name):

        df_final.write \
            .format("delta") \
            .mode('overwrite') \
                .partitionBy("Ano","Trimestre","Compania","Tipo") \
            .saveAsTable(table_name)
    else:
        df_final.createOrReplaceTempView("tmp_trafico_vias")
        spark.sql(
            f"""
                MERGE INTO {table_name} AS target
                    USING tmp_trafico_vias AS source
                    ON  target.Ano       = source.Ano
                    AND target.Trimestre   = source.Trimestre
                    AND target.Compania    = source.Compania
                    AND target.Tipo    = source.Tipo
                    WHEN MATCHED THEN
                    UPDATE SET *
                    WHEN NOT MATCHED THEN
                    INSERT *;
            """
        )
except Exception as ex:
    print('Task Fail--> ### ' + str(ex) + ' ###')
