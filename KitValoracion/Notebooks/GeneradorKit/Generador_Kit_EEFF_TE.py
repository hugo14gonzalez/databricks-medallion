# Databricks notebook source
#Hugo
%pip install --quiet openpyxl
dbutils.library.restartPython()

# COMMAND ----------

import pandas as pd
import re
import json
import shutil
import os
import unicodedata
from openpyxl import load_workbook
from copy import copy
from openpyxl.cell import MergedCell
from pyspark.sql.functions import col, regexp_replace, lower, udf
from pyspark.sql.types import StringType

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = '[["CAPEX", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/CAPEX/Capex_3Q_2024.xlsx", "CapexPorEmpresa", "CAPEX"], ["Colombia", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Colombia/Kit_Valoracion_ITCO_3Q_2024.xlsx", "ITCO (propuesta)", "EEFF_TE_Col"], ["CTEEP", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/CTEEP/CTEEP Consolidado_Kit Inversionista vISA CTEEP_3Q_2024_.xlsx", "EEFF_TE_Bra", "EEFF_TE_Bra"], ["Costera", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Costera/EEFF Costera_3Q_2024.xlsx", "EEFF_V\u00edas_Col", "EEFF_V\u00edas_Col"], ["Deuda", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx", "PerfilD", "Perfil deuda"], ["Dividendos Filiales", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Dividendos Filiales/Dividendos Filiales_3Q_2024.xlsx", "Dividendos", "Dividendos"], ["Interchile", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Interchile/CHILE_TE_3Q_2024.xlsx", "EEFF_TE_Chi", "EEFF_TE_Chi"], ["Participaciones Empresas", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Participaciones Empresas/RP_Participacion Directa & Efectiva0924_3Q_2024.xlsm", "Reporte_SP", "Participaci\u00f3n en compa\u00f1\u00edas"], ["Per\u00fa", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Per\u00fa/EEFF TE Per_3Q_2024.xlsx", "EEFF_TE_Per", "EEFF_TE_Per"], ["Proyectos", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Proyectos/Informaci\u00f3n Negocio TE_3Q_2024.xlsx", "Construcci\u00f3n proyetos", "Proyectos"], ["Transelca", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Transelca/KIT_Valoracion_Transelca_3Q_2024.xlsx", "EEFF_TE_Col", "EEFF_TE_Col"], ["Trafico v\u00edas", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Trafico v\u00edas/InfoV\u00edas-ReportesTrimestralesISA-paraRI_3Q_2024.xlsx", "Tr\u00e1fico v\u00edas", "Tr\u00e1fico v\u00edas"], ["V\u00edas Chile", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/V\u00edas Chile/EEFF V\u00edas Chile_3Q_2024.xlsx", "EEFF_V\u00edas_Chi", "EEFF_V\u00edas_Chi"], ["Vencimiento Concesiones Viales", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Vencimiento Concesiones Viales/Vencimiento Concesiones_3Q_2024.xlsx", "Vencimiento Concesiones", "Vencimiento Concesiones"], ["Deuda", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx", "KitCreditos", "Deuda consolidada cr\u00e9ditos"], ["Deuda", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx", "KitBonos", "Deuda consolidada bonos"]]'

#lectura_carpetas = '[["CAPEX", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/CAPEX/Capex_3Q_2024.xlsx", "CapexPorEmpresa", "CAPEX"], ["Colombia", "dbfs:/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Colombia/Kit_Valoracion_ITCO_3Q_2024.xlsx", "ITCO (propuesta)", "EEFF_TE_Col"]]'

full_ano = '2024'
trimestre = '3'
archivo_kit_origen = '/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Kit consolidado y separados/kit_Inversionistas ISA_Sp.xlsx'
archivo_destino = '/Volumes/kitvaloracion/base/KitValoracion/2024/3Q24/Kit consolidado y separados/kit_Inversionistas ISA_Sp_3Q24.xlsx'

# lectura_carpetas = dbutils.widgets.get('df_base')
# archivo_kit_origen = dbutils.widgets.get('ruta_kit_origen')
# archivo_destino = dbutils.widgets.get('ruta_archivo_destino')
# full_ano = dbutils.widgets.get('Anio')
# trimestre = dbutils.widgets.get('Trimestre')

ano = int(str(full_ano)[2:])

df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn('Carpeta', normalizar_udf(col('Carpeta')))\
    .filter(
        (col('Carpeta') == 'colombia') |
        (col('Carpeta') == 'costera') |
        (col('Carpeta') == 'cteep') |
        (col('Carpeta') == 'interchile') |
        (col('Carpeta') == 'peru') |
        (col('Carpeta') == 'transelca') |
        (col('Carpeta') == 'vias chile') 
    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

display(df_carpetas)

# COMMAND ----------

directorio, nombre_archivo = os.path.split(archivo_destino)

temp_path = 'tmp/'
full_temp_path = temp_path + nombre_archivo

os.makedirs(temp_path, exist_ok=True)

if not os.path.exists(full_temp_path):
    shutil.copy(archivo_kit_origen, full_temp_path)

# Ensure the destination directory exists
#directorio, nombre_archivo = os.path.split(archivo_destino)
#os.makedirs(directorio, exist_ok=True)

#if not os.path.exists(archivo_destino):
#    shutil.copy(archivo_kit_origen, archivo_destino)

# COMMAND ----------

try:
    # Ensure the directory exists before saving
    #output_dir = os.path.dirname(archivo_destino)
    #os.makedirs(output_dir, exist_ok=True)

    lectura_inicial = True
    for i in range(len(rutaArchivos)):
        print('=' * 80)
        print(f'Archivo {i+1}/{len(rutaArchivos)}: {rutaArchivos[i]}...')
        print(f'Hoja: {hojas_origen[i]}...')
        print('=' * 80)

        # Cargar los libros de Excel
        source_wb = load_workbook(rutaArchivos[i][5:], data_only=True)
        #target_wb = load_workbook(archivo_destino, data_only=True)

        if lectura_inicial:
            target_wb = load_workbook(full_temp_path, data_only=True)
            lectura_inicial = False
        else:
            target_wb = load_workbook(full_temp_path, data_only=True)

        # Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
        source_ws = source_wb[hojas_origen[i]]
        target_ws = target_wb[hojas_destino[i]]

        # hugo
        columna_encabezado_kit = 0
        columna_encabezado_kit_encontrado = False

        columna_encontrada = False
        columna_origen = 0
        dic_empresas = {}
        ultimo_titulo = ''
        ultima_fila_llena = 0

        # 1. Recorrer filas y columnas del archivo origen y hoja origen: i (archivos de cada carpeta)
        for filas in range(1,source_ws.max_row):
            for columnas in range(1,50):
                valor = str(source_ws.cell(row=filas, column=columnas).value)                

                if 'ESTADOS DE RESULTADOS' in valor \
                or 'ESTADOS SEPARADOS DE SITUACIÓN FINANCIERA' in valor\
                or 'ESTADOS DE RESULTADOS CONSOLIDADO' in valor\
                or 'ESTADOS CONSOLIDADOS DE SITUACIÓN' in valor:
                    
                    # print(f'fila: {filas}, columna: {columnas}, valor: {valor}')

                    if not columna_encontrada:
                        columna_descripcion_origen = columnas
                        for sub_fila in range(filas, filas + 6):
                            columnas_vacias = 0
                            for sub_columna in range(1, 100):
                                sub_valor = str(source_ws.cell(row=sub_fila, column=sub_columna).value) 
                                # print(f'sub_fila: {sub_fila}, sub_columna: {sub_columna}, sub_valor: {sub_valor}')
                                
                                if sub_valor == 'None':
                                    columnas_vacias = columnas_vacias + 1
                                else:
                                    columnas_vacias = 0

                                if sub_valor == f'{trimestre}T{ano}' or sub_valor == f'{trimestre}Q{ano}':
                                    columna_origen = sub_columna
                                    columna_encontrada = True
                                    break
                                
                                # Detener si hay n columnas vacías
                                if columnas_vacias >= 5:
                                    break
                            
                            if columna_encontrada:
                                break       
                    
                    #Guardar fila y columna                   
                    if valor.split('-')[-1].split()[-1] != 'ISA' and 'TRANSELCA' in valor:
                        nombre_empresa = valor.replace('ISA','').replace("  ", " ").replace(' -','-').replace('- ','-').strip()
                    else:
                        nombre_empresa = valor.replace("  ", " ").replace(' -','-').replace('- ','-').strip()

                    print(f'nombre_empresa: {nombre_empresa}')

                    if valor not in dic_empresas:
                        dic_empresas[nombre_empresa] =  {'Filas': {'inicial' : '', 'final' : ''}, 'Columna': ''}
                    
                    dic_empresas[nombre_empresa]['Filas']['inicial'] = filas
                    dic_empresas[nombre_empresa]['Columna'] = columna_origen                    
                    ultimo_titulo = nombre_empresa
                elif valor != 'None':
                    ultima_fila_llena = filas

                if ultimo_titulo != '':
                    dic_empresas[ultimo_titulo]['Filas']['final'] = ultima_fila_llena

            # if columna_origen == 0:
            #     print(f'No se encontró el valor {trimestre}T{ano}')
            #     break

        # Fin paso 1.
        print(f'dic_empresas: {dic_empresas}')
        for empresa in dic_empresas:
            if dic_empresas[empresa]['Columna'] == 0:
                print(f'no existe el valor {trimestre}T{ano} en {empresa}')

        empresa_anterior = ''
        fila_kit_empresa = 0
        dic_kit_empresas = {}

        # 2. Recorrer filas y columnas del archivo destino creado en la carpeta: temp/
        print('')
        print('-' * 80)
        print(f'Hoja destino: {hojas_destino[i]}...')
        print('-' * 80)
        for fila_kit in range(1,target_ws.max_row + 1):
            for columna_kit in range(1,10):
                valor_kit = str(target_ws.cell(row=fila_kit, column=columna_kit).value).replace("  ", " ").replace(' -','-').replace('- ','-').strip()

                if valor_kit in dic_empresas:
                    columna_descripcion_kit = columna_kit
                    
                    print(f'fila_kit: {fila_kit}, columna_kit: {columna_kit}, valor_kit: {valor_kit}')
                    print(f'columna_descripcion_kit: {columna_descripcion_kit}')

                    if valor_kit not in dic_kit_empresas:
                        dic_kit_empresas[valor_kit] =  {'Fila': '', 'Columna': ''}

                    for sub_fila_kit in range(fila_kit, fila_kit + 6):
                        sub_columna_kit_vacia = 0
                        for sub_columna_kit in range(1, 100):
                            sub_valor_kit = str(target_ws.cell(row=sub_fila_kit, column=sub_columna_kit).value)
                            
                            # print(f'sub_fila_kit: {sub_fila_kit}, sub_columna_kit: {sub_columna_kit}, sub_valor_kit: {sub_valor_kit}')

                            if bool(re.match(r"^[1-4]T\d{2}$",sub_valor_kit)) or bool(re.match(r"^[1-4]Q\d{2}$",sub_valor_kit)):
                                sub_columna_kit_vacia = 0
                                columna_encabezado_kit = sub_columna_kit + 1

                                if sub_valor_kit == f'{trimestre}T{ano}' or sub_valor_kit == f'{trimestre}Q{ano}':
                                    columna_encabezado_kit = sub_columna_kit
                                    columna_encabezado_kit_encontrado = True
                                    
                                    print(f'columna_encabezado_kit: {columna_encabezado_kit}')
                                    print(f'columna_encabezado_kit_encontrado: {columna_encabezado_kit_encontrado}')
                                    
                                    break
                            else:
                                sub_columna_kit_vacia = sub_columna_kit_vacia + 1
                            
                            if sub_columna_kit_vacia >= 5:
                                break
                
                    dic_kit_empresas[valor_kit]['Fila'] = fila_kit
                    dic_kit_empresas[valor_kit]['Columna'] = columna_encabezado_kit

                    # if empresa_anterior != '':
                    #     dic_kit_empresas[empresa_anterior]['Filas']['final'] = fila_kit - 1

                    #empresa_anterior = valor_kit

        print(f'\ndic_kit_empresas: {dic_kit_empresas}')

        # 3. Remover celdas mezcladas en archivo destino creado en la carpeta: temp/
        for rng in list(target_ws.merged_cells.ranges):
            print(f'Rango filas a remover mezcla - rng: {rng}')
            target_ws.unmerge_cells(str(rng))
        
        #se inicia el valor ultima fila procesada
        fila_maxima = 0

        # 4. Recorremos las empresas encontradas en el archivo origen y la buscamos en el archivo destino
        for empresa in dic_empresas:
            columna_origen = dic_empresas[empresa]['Columna']

            #Validamos si la empresa existe en el diccionario de las empresas encontradas en el kit
            if empresa in dic_kit_empresas:
                print(f'\nExiste la empresa {empresa} en el kit destino')

                # se inicializa las variables de filas y columnas para procesar datos
                fila_inicial_origen = dic_empresas[empresa]['Filas']['inicial']
                fila_final_origen = dic_empresas[empresa]['Filas']['final']
                fila_destino = dic_kit_empresas[empresa]["Fila"]

                #Validamos si las siguientes 6 filas de descipciones del kit existen, para saber si tiene datos o no la empresa en el kit
                descripcion_vacio = False
                celda_destino_vacias = 0
                for fila in range(fila_inicial_origen, fila_inicial_origen + 6):
                    celda_validacion = target_ws.cell(row = fila, column = columna_descripcion_kit).value
                    if celda_validacion is None:
                        celda_destino_vacias = celda_destino_vacias + 1
                    else:
                        celda_destino_vacias = 0
                    
                    if celda_destino_vacias >= 5:
                        descripcion_vacio = True

                print(f'columna_descripcion_kit: {columna_descripcion_kit}, valor: {celda_validacion}, celda_destino_vacias: {celda_destino_vacias}, descripcion_vacio: {descripcion_vacio}')

                # 4.A.1. En caso de que solo este el titulo de la empresa sin descripciones o datos. Copiar columna Descripciones
                if descripcion_vacio:
                    print(f'\nCopiando Descripciones de la columna_descripcion_origen: {columna_descripcion_origen} a columna_descripcion_kit: {columna_descripcion_kit}')
                    for fila_descripcion in range(fila_inicial_origen, fila_final_origen + 1):
                        celda_descripcion_origen = source_ws.cell(row=fila_descripcion, column=columna_descripcion_origen)
                        celda_descripcion_destino = target_ws.cell(row=fila_destino, column=columna_descripcion_kit)
                        celda_descripcion_destino_anterior = target_ws.cell(row=fila_destino, column=columna_descripcion_kit - 1)

                        # print(f'Descripcion - Fila origen: {fila_descripcion}, valor: {celda_descripcion_origen.value}. Fila destino: {fila_destino}, valor: {celda_descripcion_destino.value}')

                        #copia valor y estilos
                        celda_descripcion_destino.value = celda_descripcion_origen.value
                        celda_descripcion_destino.font = copy(celda_descripcion_destino_anterior.font)
                        celda_descripcion_destino.border = copy(celda_descripcion_destino_anterior.border)
                        celda_descripcion_destino.fill = copy(celda_descripcion_destino_anterior.fill)
                        celda_descripcion_destino.number_format = celda_descripcion_destino_anterior.number_format
                        celda_descripcion_destino.protection = copy(celda_descripcion_destino_anterior.protection)
                        celda_descripcion_destino.alignment = copy(celda_descripcion_destino_anterior.alignment)

                        #Se suma uno a la fila destino para copiar la siguiente fila
                        fila_destino = fila_destino + 1

                #se llama la variable de la columna destino y se reinicia la fila destino por si no existian descripciones
                fila_destino = dic_kit_empresas[empresa]["Fila"]
                columna_destino = dic_kit_empresas[empresa]["Columna"]

                #4.A.2. Copiar columna de valores desde fila inicial origen hasta final final origen
                print(f'\nCopiando Valores de la columna_origen: {columna_origen} a columna_destino: {columna_destino}')
                for fila_copia in range(fila_inicial_origen, fila_final_origen + 1):
                    celda_origen = source_ws.cell(row=fila_copia, column=columna_origen)
                    celda_destino = target_ws.cell(row=fila_destino, column=columna_destino)
                    celda_destino_anterior = target_ws.cell(row=fila_destino, column=columna_destino - 1)

                    # print(f'Valor - Fila origen: {fila_copia}, valor: {celda_origen.value}. Fila destino: {fila_destino}, valor: {celda_destino.value}')

                    #copia valor y estilos
                    celda_destino.value = celda_origen.value
                    celda_destino.font = copy(celda_destino_anterior.font)
                    celda_destino.border = copy(celda_destino_anterior.border)
                    celda_destino.fill = copy(celda_destino_anterior.fill)
                    celda_destino.number_format = celda_destino_anterior.number_format
                    celda_destino.protection = copy(celda_destino_anterior.protection)
                    celda_destino.alignment = copy(celda_destino_anterior.alignment)
                    
                    #Se suma uno a la fila destino para copiar la siguiente fila
                    fila_destino = fila_destino + 1

                if fila_destino > fila_maxima:
                    fila_maxima = fila_destino
            else:
                # 4.A.1. y 4.A.2.
                print(f'\nNo existe la empresa {empresa} en el kit destino')
                print(f'Copiando columna_descripcion_origen: {columna_descripcion_origen} a columna_descripcion_kit: {columna_descripcion_kit}')
                print(f'Copiando columna_origen: {columna_origen} a columna_destino: {columna_destino}')

                fila_maxima = fila_maxima + 3
                fila_origen_origen = dic_empresas[empresa]['Filas']['inicial']
                fila_origen_final = dic_empresas[empresa]['Filas']['final']

                for fila_nuevo_dato in range(fila_origen_origen, fila_origen_final + 1):
                    celda_descripcion_origen = source_ws.cell(row=fila_nuevo_dato, column=columna_descripcion_origen)
                    celda_descripcion_destino = target_ws.cell(row=fila_maxima, column=columna_descripcion_kit)
                    celda_descripcion_destino_anterior = target_ws.cell(row=fila_maxima, column=columna_descripcion_kit - 1)

                    # print(f'Descripcion - Fila origen: {fila_nuevo_dato}, valor: {celda_descripcion_origen.value}. Fila destino: {fila_maxima}, valor: {celda_descripcion_destino.value}')

                    #copia valor y estilos
                    celda_descripcion_destino.value = celda_descripcion_origen.value
                    celda_descripcion_destino.font = copy(celda_descripcion_destino_anterior.font)
                    celda_descripcion_destino.border = copy(celda_descripcion_destino_anterior.border)
                    celda_descripcion_destino.fill = copy(celda_descripcion_destino_anterior.fill)
                    celda_descripcion_destino.number_format = celda_descripcion_destino_anterior.number_format
                    celda_descripcion_destino.protection = copy(celda_descripcion_destino_anterior.protection)
                    celda_descripcion_destino.alignment = copy(celda_descripcion_destino_anterior.alignment)

                    celda_origen = source_ws.cell(row=fila_nuevo_dato, column=columna_origen)
                    celda_destino = target_ws.cell(row=fila_maxima, column=columna_destino)
                    celda_destino_anterior = target_ws.cell(row=fila_maxima, column=columna_destino - 1)

                    # print(f'Valor - Fila origen: {fila_nuevo_dato}, valor: {celda_origen.value}. Fila destino: {fila_maxima}, valor: {celda_destino.value}')

                    #copia valor y estilos
                    celda_destino.value = celda_origen.value
                    celda_destino.font = copy(celda_destino_anterior.font)
                    celda_destino.border = copy(celda_destino_anterior.border)
                    celda_destino.fill = copy(celda_destino_anterior.fill)
                    celda_destino.number_format = celda_destino_anterior.number_format
                    celda_destino.protection = copy(celda_destino_anterior.protection)
                    celda_destino.alignment = copy(celda_destino_anterior.alignment)

                    fila_maxima = fila_maxima + 1
        
        print(f'Guardar y copiar en archivo destino: {archivo_destino}')
        target_wb.save(full_temp_path)
        shutil.copy(full_temp_path, archivo_destino)
except Exception as ex:
    print('Task Fail--> ### ' + str(ex) + ' ###')
