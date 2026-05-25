# Databricks notebook source
from openpyxl import load_workbook
import re
from copy import copy
import json
from pyspark.sql.functions import col

# COMMAND ----------

# rutaArchivos = ['absf:/Volumes/test_data/base/4q23/Trafico Vías/InfoVías-ReportesTrimestralesISA-paraRI_3Q_2024.xlsx']
# hojas_origen = ['Tráfico vías']

# ruta_destino = '/Volumes/test_data/base/4q24/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx'
# hojas_destino = ['Tráfico vías']

# nombre_archivo_final = ruta_destino

# full_ano = 2024
# trimestre = 3
# ano = int(str(full_ano)[2:])

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .filter(
        (col('Carpeta') == 'Trafico Vías')
    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

archivo_destino = dbutils.widgets.get('ruta_archivo_destino')

#dbutils.fs.cp(archivo_destino, "/Workspace/Users/itco_e_jacevedo@intercolombia.com/Procesamiento/Generador Kit/prueba_kit.xlsx")

full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
ano = int(str(full_ano)[2:])

nombre_archivo_final = 'prueba_all_kit.xlsx'

# COMMAND ----------

# Cargar los libros de Excel
source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
target_wb = load_workbook(nombre_archivo_final, data_only=True)

# Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
source_ws = source_wb[hojas_origen[0]]
target_ws = target_wb[hojas_destino[0]]

# COMMAND ----------

fila_encabezado_origen = 0
columna_encabezado_origen = 0
valor_encontrado = False
columna_descripcion_origen = 0

dic_rutas = {}

for fila in range(2, source_ws.max_row+1):
    for columna in range(1, source_ws.max_column+1):

        valor = str(source_ws.cell(row=fila, column=columna).value)
        #print(valor)

        if (bool(re.match(r"^[1-4]T\d{2}$",valor)) or bool(re.match(r"^[1-4]Q\d{2}$",valor))) and not valor_encontrado:
            #print(f'{valor} -- {trimestre}Q{ano}')
            fila_encabezado_origen = fila
            if valor == f"{trimestre}T{ano}" or valor == f"{trimestre}Q{ano}":
                columna_encabezado_origen = columna
                valor_encontrado = True

        elif 'RUTA' in valor:
            dic_rutas[valor] = {'fila': fila}
            columna_descripcion_origen = columna
            

if fila_encabezado_origen == 0 or columna_encabezado_origen == 0:
    print('no se encontró el trimestre, no puede continuar')
else:
    print(f'fila {fila_encabezado_origen} - columna {columna_encabezado_origen}')
    print(dic_rutas)


# COMMAND ----------

fila_encabezado_destino = 0
columna_encabezado_destino = 0
columna_nombres_destino = 0
valor_encontrado = False

dic_rutas_destino = {}

for fila in range(2, target_ws.max_row+1):
    for columna in range(1, target_ws.max_column+1):

        valor = str(target_ws.cell(row=fila, column=columna).value)
        #print(valor)

        if (bool(re.match(r"^[1-4]T\d{2}$",valor)) or bool(re.match(r"^[1-4]Q\d{2}$",valor)) or bool(re.match(r"^\d{4}-[1-4]Q$",valor)) or bool(re.match(r"^\d{4}-[1-4]Q$",valor))) and not valor_encontrado:
            #print(f'{valor} -- {trimestre}Q{ano}')

            fila_encabezado_destino = fila

            if valor == f"{trimestre}T{ano}" or valor == f"{trimestre}Q{ano}":
                columna_encabezado_destino = columna
                valor_encontrado = True
            
        elif 'RUTA' in valor:
            dic_rutas_destino[valor] = {'fila': fila}
            columna_nombres_destino = columna

if columna_encabezado_destino == 0:
    columna_encabezado_destino = columna + 1

#print(f'fila {fila_encabezado_destino} - columna {columna_encabezado_destino}')
# print(f'Columna nombres: {columna_nombres_destino}')
dic_rutas_destino

# COMMAND ----------

fila_encabezado_origen
columna_encabezado_origen

fila_encabezado_destino
columna_encabezado_destino


celda_encabezado_origen = source_ws.cell(row=fila_encabezado_origen, column=columna_encabezado_origen)
celda_encabezado_destino = target_ws.cell(row=fila_encabezado_destino, column=columna_encabezado_destino)

celda_encabezado_destino.value = celda_encabezado_origen.value

celda_encabezado_destino.font = copy(celda_encabezado_origen.font)
celda_encabezado_destino.border = copy(celda_encabezado_origen.border)
celda_encabezado_destino.fill = copy(celda_encabezado_origen.fill)
celda_encabezado_destino.number_format = celda_encabezado_origen.number_format
celda_encabezado_destino.protection = copy(celda_encabezado_origen.protection)
celda_encabezado_destino.alignment = copy(celda_encabezado_origen.alignment)

# COMMAND ----------

for ruta in dic_rutas:
    if ruta in dic_rutas_destino:
        # print(ruta)
        # print()
        i = 0
        celda_con_data = True
        while celda_con_data:
            
            celda_descripcion_origen = source_ws.cell(row=dic_rutas[ruta]['fila'] + i, column=columna_descripcion_origen)

            print(f'{celda_descripcion_origen.value}')

            if str(celda_descripcion_origen.value) == 'None' or re.fullmatch(r'\s+', str(celda_descripcion_origen.value)):
                celda_con_data = False
                break

            celda_origen = source_ws.cell(row=dic_rutas[ruta]['fila'] + i, column=columna_encabezado_origen)
            celda_destino = target_ws.cell(row=dic_rutas_destino[ruta]['fila'] + i, column=columna_encabezado_destino)

            celda_destino.value = celda_origen.value

            celda_destino.font = copy(celda_origen.font)
            celda_destino.border = copy(celda_origen.border)
            celda_destino.fill = copy(celda_origen.fill)
            celda_destino.number_format = celda_origen.number_format
            celda_destino.protection = copy(celda_origen.protection)
            celda_destino.alignment = copy(celda_origen.alignment)

        
        

            i = i + 1
        print('*' * 40)
        ultima_ruta = ruta
    else:
        #print(f'No se encontró {ruta}')

        ultima_fila_procesada = dic_rutas_destino[ultima_ruta]['fila'] + 4

        target_ws.insert_rows(ultima_fila_procesada, 4)

        i = 0
        celda_con_data = True
        while celda_con_data:

            celda_descripcion_origen = source_ws.cell(row=dic_rutas[ruta]['fila'] + i, column=columna_descripcion_origen)

            #print(f'{celda_descripcion_origen.value}')

            if str(celda_descripcion_origen.value) == 'None' or re.fullmatch(r'\s+', str(celda_descripcion_origen.value)):
                celda_con_data = False
                break

            celda_descripcion_destino = target_ws.cell(row=ultima_fila_procesada + i, column=columna_nombres_destino)

            celda_descripcion_destino.value = celda_descripcion_origen.value

            celda_descripcion_destino.font = copy(celda_descripcion_origen.font)
            celda_descripcion_destino.border = copy(celda_descripcion_origen.border)
            celda_descripcion_destino.fill = copy(celda_descripcion_origen.fill)
            celda_descripcion_destino.number_format = celda_descripcion_origen.number_format
            celda_descripcion_destino.protection = copy(celda_descripcion_origen.protection)
            celda_descripcion_destino.alignment = copy(celda_descripcion_origen.alignment)

            celda_origen = source_ws.cell(row=dic_rutas[ruta]['fila'] + i, column=columna_encabezado_origen)
            celda_destino = target_ws.cell(row=ultima_fila_procesada + i, column=columna_encabezado_destino)

            celda_destino.value = celda_origen.value

            celda_destino.font = copy(celda_origen.font)
            celda_destino.border = copy(celda_origen.border)
            celda_destino.fill = copy(celda_origen.fill)
            celda_destino.number_format = celda_origen.number_format
            celda_destino.protection = copy(celda_origen.protection)
            celda_destino.alignment = copy(celda_origen.alignment)


            i = i + 1

            
        #print('*' * 40)

        dic_rutas_destino = {
            ruta: {'fila': valor['fila'] + 4}
            for ruta, valor in dic_rutas_destino.items()
        }
    

# COMMAND ----------

#Guardar los cambios en el archivo destino
with open(nombre_archivo_final, 'wb') as f:
    target_wb.save(f)
