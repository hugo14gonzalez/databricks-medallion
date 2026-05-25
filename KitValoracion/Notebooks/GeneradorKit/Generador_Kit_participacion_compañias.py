# Databricks notebook source
import json
import os
import unicodedata
import shutil
from openpyxl import load_workbook
from copy import copy
from pyspark.sql.functions import col, regexp_replace, lower, udf
from pyspark.sql.types import StringType


# COMMAND ----------


# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
ano = int(str(full_ano)[2:])

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn(
        'Carpeta',
        normalizar_udf(col('Carpeta'))
    )\
    .filter(
        (col('Carpeta') == 'participaciones empresas') 

    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

#archivo_kit_origen = dbutils.widgets.get('ruta_kit_origen')
archivo_destino = dbutils.widgets.get('ruta_archivo_destino')
directorio, nombre_archivo = os.path.split(archivo_destino)
temp_path = 'tmp/'
full_temp_path = temp_path + nombre_archivo
shutil.copy(archivo_destino,full_temp_path)


# COMMAND ----------

# Cargar los libros de Excel
source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
target_wb = load_workbook(full_temp_path, data_only=True)

# Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
source_ws = source_wb[hojas_origen[0]]
target_ws = target_wb[hojas_destino[0]]

# COMMAND ----------

for rng in list(target_ws.merged_cells.ranges):
    target_ws.unmerge_cells(str(rng))

# COMMAND ----------

#copiar primer cuadro
fila_vacia = False
fila_sin_datos = 1
salto = False

for fila in range(1, source_ws.max_row + 1):
    columna_vacia = 0
    for columna in range(1, 12):

        valor = str(source_ws.cell(row=fila, column=columna).value)

        if valor != 'None':
            fila_vacia = False
            fila_sin_datos = 0

            if salto:
                celda_destino = target_ws.cell(row=fila + 1, column= columna - 1)
            else:
                celda_destino = target_ws.cell(row=fila + 3, column= columna)

            celda_destino.value = valor  
        else:
            columna_vacia += 1
        
        # print(f'fila: {fila}, columna: {columna}')
        # print(valor)
        
        if columna_vacia >= 11:
            fila_vacia = True
            
            break 
    
    if fila_vacia:
        fila_sin_datos += 1

    if fila_sin_datos >= 3:
        salto = True
    
    # print(fila_sin_datos)
    # print('*' * 40)
    


# COMMAND ----------

#Guardar los cambios en el archivo destino
# with open(archivo_destino, 'wb') as f:
#     wb_destino.save(f)

target_wb.save(full_temp_path)
shutil.copy(full_temp_path,archivo_destino)
