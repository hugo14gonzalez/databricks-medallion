# Databricks notebook source
import os
import json
import pandas as pd
import unicodedata
from openpyxl import load_workbook
from copy import copy
from pyspark.sql.functions import col, regexp_replace, lower, udf
from databricks.sdk.runtime import *
from pyspark.sql.types import StringType

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

rutaOrigen = ['abfs:/Volumes/test_data/base/4q23/CAPEX/Capex_3Q_2024Kit.xlsx']
hojas_origen = ['CapexPorEmpresa']

rutaDestino = ['abfs:/Volumes/test_data/base/4q23/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx']
hojas_destino = ['CAPEX']

full_ano = 2025

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))

display(df_carpetas)

df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn(
        'Carpeta',
        normalizar_udf(col('Carpeta'))
    )\
    .filter(
        (col('Carpeta') == 'capex') 

    )

display(df_carpetas)

rutaOrigen = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

archivo_kit_origen = dbutils.widgets.get('ruta_kit_origen')
archivo_destino = dbutils.widgets.get('ruta_archivo_destino')


full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
ano = int(str(full_ano)[2:])

# COMMAND ----------

#Cargar los libros
wb_origen = load_workbook(rutaOrigen[0][5:], data_only=True)
wb_destino = load_workbook(archivo_destino)

#Seleccionar las hojas
ws_origen = wb_origen[hojas_origen[0]]
ws_destino = wb_destino[hojas_destino[0]]


# COMMAND ----------

fila_inicio_descripcion_origen = 0
columna_inicio_descripcion_origen = 0
fila_inicio_datos_origen = 0
columna_inicio_datos_origen = 0

result = ''

for fila in range(1, ws_origen.max_row + 1):
    for columna in range(1, ws_origen.max_column + 1):
        celda = ws_origen.cell(row=fila, column=columna)
        valor = str(celda.value)
        if valor is not None:
            #Buscar la ubicación donde está el titulo de capex consolidado
            if valor.lower() == 'capex consolidado':
                fila_inicio_descripcion_origen = fila
                columna_inicio_descripcion_origen = columna
            
            if valor == str(full_ano):
                fila_inicio_datos_origen = fila
                columna_inicio_datos_origen = columna

if fila_inicio_descripcion_origen == 0\
or columna_inicio_descripcion_origen == 0\
or fila_inicio_datos_origen == 0\
or columna_inicio_datos_origen == 0:
    dbutils.notebook.exit("Error: No se encontró el periodo en el archivo origen")
else:
    print(f'fila inicio descripcion: {fila_inicio_descripcion_origen} - columna inicio datos: {columna_inicio_descripcion_origen}')
    print(f'fila inicio datos: {fila_inicio_datos_origen} - columna inicio datos: {columna_inicio_datos_origen}')


# COMMAND ----------

print(ws_origen.cell(row=4, column=12).value)

# COMMAND ----------

dic_pais_origen = {}
pais = ''
for fila_datos in range(fila_inicio_descripcion_origen + 1, ws_origen.max_row ):

    valor_anterior = ws_origen.cell(row=fila_datos - 1 , column=columna_inicio_descripcion_origen).value
    celda = ws_origen.cell(row=fila_datos, column=columna_inicio_descripcion_origen)
    valor = celda.value
    valor_siguiente = ws_origen.cell(row=fila_datos + 1 , column=columna_inicio_descripcion_origen).value

    if valor_anterior is None:
        pais = valor
        dic_pais_origen[pais] = {}

    elif valor is not None:
        if celda.alignment.indent == 1:
            dic_pais_origen[pais][valor] = fila_datos

if dic_pais_origen == {}:
    dbutils.notebook.exit("Error: Fallo en el archivo origen, al consolidar los datos")
else:
    print(dic_pais_origen)

# COMMAND ----------


fila_descripcion_inicio_destino = 0
columna_descripcion_inicio_destino = 0
fila_inicio_datos_destino = 0
columna_inicio_datos_destino = 0

for fila in range(1, ws_destino.max_row + 1):
    for columna in range(1, ws_destino.max_column + 1):
        celda = ws_destino.cell(row=fila, column=columna)
        valor = str(celda.value)

        if valor is not None:

            if valor.lower() == 'capex consolidado':
                fila_descripcion_inicio_destino = fila
                columna_descripcion_inicio_destino = columna

            if valor == str(full_ano):
                fila_inicio_datos_destino = fila
                columna_inicio_datos_destino = columna
                break

#print(f'fila descripcion: {fila_descripcion_inicio_destino} - columna descripcion: {columna_descripcion_inicio_destino}')
    
if fila_descripcion_inicio_destino == 0\
or columna_descripcion_inicio_destino == 0\
or fila_inicio_datos_destino == 0\
or columna_inicio_datos_destino == 0:
    dbutils.notebook.exit("Error: fallo al ubicar la celda en el archivo destino")
else:
    print(f'fila descripcion: {fila_descripcion_inicio_destino} - columna descripcion: {columna_descripcion_inicio_destino}')
    print(f'fila inicio datos: {fila_inicio_datos_destino} - columna inicio datos: {columna_inicio_datos_destino}')
   

# COMMAND ----------

dic_pais_destino = {}
pais = ''
for fila_datos in range(fila_descripcion_inicio_destino + 1, ws_destino.max_row):

    valor_anterior = ws_destino.cell(row=fila_datos - 1 , column=columna_descripcion_inicio_destino).value
    celda = ws_destino.cell(row=fila_datos, column=columna_descripcion_inicio_destino)
    valor = celda.value
    valor_siguiente = ws_destino.cell(row=fila_datos + 1 , column=columna_descripcion_inicio_destino).value

    #print(valor)
    if valor_anterior is None:
        pais = valor
        dic_pais_destino[pais] = {}

    elif valor is not None:
        if celda.alignment.indent == 1:
            dic_pais_destino[pais][valor] = fila_datos

if dic_pais_destino == {}:
    dbutils.notebook.exit("Error: fallo al ubicar la celda en el archivo destino")
else:
    print(dic_pais_destino)

# COMMAND ----------

lista_filas_pais = []

for pais, empresas in dic_pais_destino.items():
    primera_fila = True
    for empresa,fila in empresas.items():
        if primera_fila:
            lista_filas_pais.append(fila - 1)
            primera_fila = False
        ultima_fila = fila

    lista_filas_pais.append(ultima_fila + 1)


print(lista_filas_pais)

# COMMAND ----------

col_destino = columna_inicio_datos_destino - 1
fila_pais = 0

for columna in range(columna_inicio_datos_origen,ws_origen.max_column):

    celda_destino = ws_destino.cell(row=fila_descripcion_inicio_destino, column=col_destino)
    celda_destino_anterior = ws_destino.cell(row=fila_descripcion_inicio_destino, column=col_destino - 1)

    celda_destino.value = ws_origen.cell(row=fila_inicio_descripcion_origen, column=columna - 1).value

    celda_destino.font = copy(celda_destino_anterior.font)
    celda_destino.border = copy(celda_destino_anterior.border)
    celda_destino.fill = copy(celda_destino_anterior.fill)
    celda_destino.number_format = celda_destino_anterior.number_format
    celda_destino.alignment = copy(celda_destino_anterior.alignment)

    
    for fila in lista_filas_pais: 
        celda_pais = ws_destino.cell(row=fila, column=col_destino)
        celda_pais_anterior = ws_destino.cell(row=fila, column=col_destino - 1)

        celda_pais.font = copy(celda_pais_anterior.font)
        celda_pais.border = copy(celda_pais_anterior.border)
        celda_pais.fill = copy(celda_pais_anterior.fill)
        celda_pais.number_format = celda_pais_anterior.number_format
        celda_pais.alignment = copy(celda_pais_anterior.alignment)



    col_destino = col_destino + 1

# COMMAND ----------

for pais, dic_filas in dic_pais_origen.items():
    if pais in dic_pais_destino:
        #print(pais)
        #print('*' * 40)
        for empresa, fila in dic_filas.items():
            #print(empresa)
            #print('-' * 40)
            if empresa in dic_pais_destino[pais]:
                columna_copia = columna_inicio_datos_destino

                for columna_datos_origen in range(columna_inicio_datos_origen, ws_origen.max_column - 1):
                
                    celda_origen = ws_origen.cell(row=fila, column=columna_datos_origen)
                    celda_destino = ws_destino.cell(row=dic_pais_destino[pais][empresa], column=columna_copia)
                    celda_destino_anterior = ws_destino.cell(row=dic_pais_destino[pais][empresa], column=columna_copia - 1)

                    celda_destino.value = celda_origen.value

                    # print(f'{fila_inicio_datos_destino}.{columna_copia}) {ws_destino.cell(row=fila_inicio_datos_destino, column=columna_copia).value}')
                    # print('*' * 20)

                    celda_destino.font = copy(celda_destino_anterior.font)
                    celda_destino.border = copy(celda_destino_anterior.border)
                    celda_destino.fill = copy(celda_destino_anterior.fill)
                    celda_destino.number_format = celda_destino_anterior.number_format
                    #celda_destino.protection = copy(celda_origen.protection)
                    celda_destino.alignment = copy(celda_destino_anterior.alignment)

                    ultima_fila_procesada = dic_pais_destino[pais][empresa]

                    columna_copia = columna_copia + 1
            else:
                for i in range(len(list(dic_pais_destino[pais]))):
                    if list(dic_pais_origen[pais])[i] != list(dic_pais_destino[pais])[i]:
                        # print([list(dic_pais_destino[pais])[i]])
                        # print(dic_pais_destino[pais][list(dic_pais_destino[pais])[i]])
                        fila_copia = dic_pais_destino[pais][list(dic_pais_destino[pais])[i]]
                        ws_destino.insert_rows(idx=fila_copia, amount=1)
                        break
                   

                celda_descripcion_origen = ws_origen.cell(row=dic_pais_origen[pais][empresa], column=columna_inicio_descripcion_origen)
                celda_descripcion_destino = ws_destino.cell(row=fila_copia, column=columna_descripcion_inicio_destino)
                celda_descripcion_destino_formato = ws_destino.cell(row=fila_copia + 1, column=columna_descripcion_inicio_destino)

                celda_descripcion_destino.value = celda_descripcion_origen.value

                celda_descripcion_destino.font = copy(celda_descripcion_destino_formato.font)
                celda_descripcion_destino.border = copy(celda_descripcion_destino_formato.border)
                celda_descripcion_destino.fill = copy(celda_descripcion_destino_formato.fill)
                celda_descripcion_destino.number_format = celda_descripcion_destino_formato.number_format
                celda_descripcion_destino.alignment = copy(celda_descripcion_destino_formato.alignment)

                columna_copia = columna_inicio_datos_destino

                for columna_datos_origen in range(columna_inicio_datos_origen, ws_origen.max_column - 1):

                    celda_dato_origen = ws_origen.cell(row=dic_pais_origen[pais][empresa], column=columna_datos_origen)
                    celda_dato_destino = ws_destino.cell(row=fila_copia, column=columna_copia)
                    celda_dato_destino_formato = ws_destino.cell(row=fila_copia + 1, column=columna_copia)

                    celda_dato_destino.value = celda_dato_origen.value

                    celda_dato_destino.font = copy(celda_dato_destino_formato.font)
                    celda_dato_destino.border = copy(celda_dato_destino_formato.border)
                    celda_dato_destino.fill = copy(celda_dato_destino_formato.fill)
                    celda_dato_destino.number_format = celda_dato_destino_formato.number_format
                    celda_dato_destino.alignment = copy(celda_dato_destino_formato.alignment)

                    columna_copia = columna_copia + 1
                

                dic_pais_destino = {
                    pais : {empresa: fila + 1 for empresa, fila in dic_filas.items() }
                    for pais, dic_filas in dic_pais_destino.items()
                }





# COMMAND ----------

# with open(archivo_destino, 'wb') as f:
#     wb_destino.save(f)

wb_destino.save(archivo_destino)

# COMMAND ----------

del_variables = [var for var in locals() if not var.startswith('__')]
for var in del_variables:
    del locals()[var]

var = None
