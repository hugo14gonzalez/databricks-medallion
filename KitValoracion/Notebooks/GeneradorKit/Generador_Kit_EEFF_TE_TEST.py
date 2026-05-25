# Databricks notebook source
from openpyxl import load_workbook, Workbook
import pandas as pd
from copy import copy
import re
from openpyxl.cell import MergedCell
import json
from pyspark.sql.functions import col
import shutil
import os

# COMMAND ----------

dic_lectura_archivo = {}

dic_lectura_archivo['Origen'] = {
    'Colombia':
        {
            'Archivo': '/Volumes/test_data/base/4q24/Colombia/Kit_Valoracion_Q4_ITCO.xlsx',
            'Hoja Origen': 'ITCO (propuesta)',
            'Hoja Destino': 'EEFF_TE_Col'
        },
    'Costera':
        {
            'Archivo':  '/Volumes/test_data/base/4q24/Costera/EEFF Costera.xlsx',
            'Hoja Origen':  'EEFF_Vías_Col',
            'Hoja Destino': 'EEFF_Vías_Col'
        },
    
    'CTEEP':
        {
            'Archivo': '/Volumes/test_data/base/4q24/CTEEP/2024/CTEEP Consolidado_Kit Inversionista vISA CTEEP 3Q24_.xlsx',
            'Hoja Origen': 'EEFF_TE_Bra',
            'Hoja Destino': 'EEFF_TE_Bra'
        },
    'Interchile':
        {
            'Archivo':  '/Volumes/test_data/base/4q24/Interchile/CHILE_TE_3T23.xlsx',
            'Hoja Origen': 'EEFF_TE_Chi',
            'Hoja Destino': 'EEFF_TE_Chi'
        },
    'Perú':
        {
            'Archivo': '/Volumes/test_data/base/4q24/Perú/EEFF TE Per 2023 T3.xlsx',
            'Hoja Origen': 'EEFF_TE_Per',
            'Hoja Destino': 'EEFF_TE_Per'
        },
    'Transelca':
        {
            'Archivo': '/Volumes/test_data/base/4q24/Transelca/KIT_Valoracion_Q4_Transelca 2024.xlsx',
            'Hoja Origen': 'EEFF_TE_Col',
            'Hoja Destino': 'EEFF_TE_Col'
        },
    'Vías Chile':
        {
            'Archivo': '/Volumes/test_data/base/4q24/Vías Chile/EEFF Vías Chile 4T2023.xlsx',
            'Hoja Origen': 'EEFF_Vías_Chi',
            'Hoja Destino': 'EEFF_Vías_Chi'
        }
}

dic_lectura_archivo['Destino'] = '/Volumes/test_data/base/4q24/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx'



print(dic_lectura_archivo)

# COMMAND ----------

df_carpetas = spark.createDataFrame(
    [(key, value['Archivo'], value['Hoja Origen'], value['Hoja Destino']) for key, value in dic_lectura_archivo['Origen'].items()],
    ['Pais', 'RutaArchivo', 'HojaOrigen', 'HojaDestino']
)

display(df_carpetas)

# COMMAND ----------


rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

archivo_kit_origen = '/Volumes/test_data/base/4q23/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx'
archivo_destino = '/Workspace/Users/itco_e_jacevedo@intercolombia.com/prueba_kit.xlsx'

full_ano = 2024
trimestre = 3
ano = int(str(full_ano)[2:])

#display(df_carpetas)

# COMMAND ----------


if not os.path.exists(archivo_destino):

    # Obtener el nombre del archivo y la ruta del directorio
    directorio, nombre_archivo = os.path.split(archivo_destino)

    # Crear la ruta del archivo antiguo con el guion bajo
    ruta_archivo_antiguo = os.path.join(directorio, f"{nombre_archivo}")

    # Copiar el archivo antiguo al nuevo archivo con el nombre original
    shutil.copy(archivo_kit_origen, ruta_archivo_antiguo)


# COMMAND ----------

try:


    lectura_inicial = True
    for i in range(len(rutaArchivos)):

        # Cargar los libros de Excel
        source_wb = load_workbook(rutaArchivos[i], data_only=True)
        #target_wb = load_workbook(archivo_destino, data_only=True)

        if lectura_inicial:
            target_wb = load_workbook(archivo_destino, data_only=True)
            lectura_inicial = False
        else:
            target_wb = load_workbook(archivo_destino, data_only=True)

        # Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
        source_ws = source_wb[hojas_origen[i]]
        target_ws = target_wb[hojas_destino[i]]

        columna_encontrada = False
        columna_origen = 0
        dic_empresas = {}
        ultimo_titulo = ''
        ultima_fila_llena = 0

        for filas in range(1,source_ws.max_row):
            for columnas in range(1,50):
                valor = str(source_ws.cell(row=filas, column=columnas).value)
                
                # print('*' * 40)
                # print(f'fila: {filas}, columna {columnas}')
                # print(valor)

                if 'ESTADOS DE RESULTADOS' in valor \
                or 'ESTADOS SEPARADOS DE SITUACIÓN FINANCIERA' in valor\
                or 'ESTADOS DE RESULTADOS CONSOLIDADO' in valor\
                or 'ESTADOS CONSOLIDADOS DE SITUACIÓN' in valor:

                    #print(valor)
                    if not columna_encontrada:
                        columna_descripcion_origen = columnas
                        for sub_fila in range(filas, filas + 6):
                            columnas_vacias = 0
                            for sub_columna in range(1, 100):

                                sub_valor = str(source_ws.cell(row=sub_fila, column=sub_columna).value) 

                                # print('*' * 40)
                                # print(f'fila: {sub_fila}, columna: {sub_columna}')
                                # print(sub_valor)
                                
                                if sub_valor == 'None':
                                    columnas_vacias = columnas_vacias + 1
                                else:
                                    columnas_vacias = 0

                                if sub_valor == f'{trimestre}T{ano}' or sub_valor == f'{trimestre}Q{ano}':
                                    columna_origen = sub_columna
                                    columna_encontrada = True
                                    break
                                
                                # print(f'vacios: {columnas_vacias}')
                                # print('*' * 40)

                                if columnas_vacias >= 5:
                                    break
                            
                            if columna_encontrada:
                                break       
                    
                    #Guardar fila y columna
                    


                    if valor.split('-')[-1].split()[-1] != 'ISA' and 'TRANSELCA' in valor:
                        nombre_empresa = valor.replace('ISA','').replace("  ", " ").replace(' -','-').replace('- ','-').strip()
                    else:
                        nombre_empresa = valor.replace("  ", " ").replace(' -','-').replace('- ','-').strip()


                    #nombre_empresa = valor.replace("  ", " ").replace(' -','-').replace('- ','-').strip()

                    # print(f'{filas}.{columnas}) {nombre_empresa}')

                    if valor not in dic_empresas:
                        dic_empresas[nombre_empresa] =  {'Filas': {'inicial' : '', 'final' : ''}, 'Columna': ''}
                    
                    dic_empresas[nombre_empresa]['Filas']['inicial'] = filas
                    dic_empresas[nombre_empresa]['Columna'] = columna_origen
                    
                    ultimo_titulo = nombre_empresa

                elif valor != 'None':
                    ultima_fila_llena = filas

                if ultimo_titulo != '':
                    dic_empresas[ultimo_titulo]['Filas']['final'] = ultima_fila_llena

            # if columna_origen == 0:
            #     print(f'No se encontró el valor {trimestre}T{ano}')
            #     break

        for empresa in dic_empresas:
            if dic_empresas[empresa]['Columna'] == 0:
                print(f'no existe el valor {trimestre}T{ano} en {empresa}')


        empresa_anterior = ''
        fila_kit_empresa = 0
        dic_kit_empresas = {}

        for fila_kit in range(1,target_ws.max_row + 1):

            for columna_kit in range(1,10):
                valor_kit = str(target_ws.cell(row=fila_kit, column=columna_kit).value).replace("  ", " ").replace(' -','-').replace('- ','-').strip()

                # print(fila_kit)

                # print(f'Fila: {fila_kit}, Columna: {columna_kit}')
                # print(valor_kit)
                # print('*' * 40)

                if valor_kit in dic_empresas:

                    columna_descripcion_kit = columna_kit
                    # print(f'Fila: {fila_kit}, Columna: {columna_kit}')
                    # print(valor_kit)
                    # print('*' * 40)

                    if valor_kit not in dic_kit_empresas:
                        dic_kit_empresas[valor_kit] =  {'Fila': '', 'Columna': ''}

                    for sub_fila_kit in range(fila_kit, fila_kit + 6):
                        sub_columna_kit_vacia = 0
                        for sub_columna_kit in range(1, 100):
                            sub_valor_kit = str(target_ws.cell(row=sub_fila_kit, column=sub_columna_kit).value)

                            # print(f'Fila: {sub_fila_kit}, Columna: {sub_columna_kit}')
                            # print(sub_valor_kit)
                            # print(f'Vacia: {sub_columna_kit_vacia}')
                            # print(f'*' * 40)

                            if bool(re.match(r"^[1-4]T\d{2}$",sub_valor_kit)) or bool(re.match(r"^[1-4]Q\d{2}$",sub_valor_kit)):
                                sub_columna_kit_vacia = 0
                                columna_encabezado_kit = sub_columna_kit + 1

                                # print(f'Fila: {sub_fila_kit}, Columna: {sub_columna_kit}')
                                # print(sub_valor_kit)
                                # print(sub_columna_kit_vacia)
                                # print(f'*' * 40)

                                if sub_valor_kit == f'{trimestre}T{ano}' or sub_valor_kit == f'{trimestre}Q{ano}':
                                    
                                    columna_encabezado_kit = sub_columna_kit
                                    columna_encabezado_kit_encontrado = True
                                    break
                            else:
                                sub_columna_kit_vacia = sub_columna_kit_vacia + 1
                            
                            if sub_columna_kit_vacia >= 5:
                                break
                
                    #fila_kit_empresa = fila_kit
                    #nombre_empresa = valor
                    #print(dic_kit_empresas)

                    dic_kit_empresas[valor_kit]['Fila'] = fila_kit
                    dic_kit_empresas[valor_kit]['Columna'] = columna_encabezado_kit

                    # if empresa_anterior != '':
                    #     dic_kit_empresas[empresa_anterior]['Filas']['final'] = fila_kit - 1

                    #empresa_anterior = valor_kit

            # print('-' * 40)
            # print(columna_encabezado_kit)
            #print(dic_empresas)

        for rng in list(target_ws.merged_cells.ranges):
            target_ws.unmerge_cells(str(rng))

        
        #se inicia el valor ultima fila procesada
        fila_maxima = 0

        #Recorremos las empresas encontradas en el archivo origen
        for empresa in dic_empresas:

            columna_origen = dic_empresas[empresa]['Columna']

            #Validamos si la empresa existe en el diccionario de las empresas encontradas en el kit
            if empresa in dic_kit_empresas:

                # se inicializa las variables de filas y columnas para procesar datos
                fila_inicial_origen = dic_empresas[empresa]['Filas']['inicial']
                fila_final_origen = dic_empresas[empresa]['Filas']['final']
                fila_destino = dic_kit_empresas[empresa]["Fila"]

                #Validamos si las siguientes 6 filas de descipciones del kit existen, para saber si tiene datos o no la empresa en el kit
                #filas_vacias = 0
                descripcion_vacio = False
                celda_destino_vacias = 0
                for fila in range(fila_inicial_origen, fila_inicial_origen + 6):
                    celda_validacion = target_ws.cell(row = fila, column = columna_descripcion_kit).value
                    if celda_validacion is None:
                        celda_destino_vacias = celda_destino_vacias + 1
                    else:
                        celda_destino_vacias = 0
                    
                    if celda_destino_vacias >= 5:
                        descripcion_vacio = True
                
                #En caso de que solo este el titulo de la empresa sin descripciones o datos se empieza el proceso
                if descripcion_vacio:

                    #Recorremos las filas de descripciones del origen, para copiarlas en el destino
                    for fila_descripcion in range(fila_inicial_origen, fila_final_origen + 1):

                        # se ubican las celdas de origen y destino de las descripciones
                        celda_descripcion_origen = source_ws.cell(row=fila_descripcion, column=columna_descripcion_origen)
                        celda_descripcion_destino = target_ws.cell(row=fila_destino, column=columna_descripcion_kit)
                        celda_descripcion_destino_anterior = target_ws.cell(row=fila_destino, column=columna_descripcion_kit - 1)

                        # se copian los valores de la celda de descripcion del origen a la celda de descripcion del destino
                        celda_descripcion_destino.value = celda_descripcion_origen.value

                        #se copian los estilos de la celda del archivo origen al destino
                        celda_descripcion_destino.font = copy(celda_descripcion_destino_anterior.font)
                        celda_descripcion_destino.border = copy(celda_descripcion_destino_anterior.border)
                        celda_descripcion_destino.fill = copy(celda_descripcion_destino_anterior.fill)
                        celda_descripcion_destino.number_format = celda_descripcion_destino_anterior.number_format
                        celda_descripcion_destino.protection = copy(celda_descripcion_destino_anterior.protection)
                        celda_descripcion_destino.alignment = copy(celda_descripcion_destino_anterior.alignment)

                        #Se le suma uno a la fila destino, para que en archivo destino copie en la siguiente fila
                        fila_destino = fila_destino + 1

                #se llama la variable de la columna destino y se reinicia la fila destino por si no existian descripciones
                fila_destino = dic_kit_empresas[empresa]["Fila"]
                columna_destino = dic_kit_empresas[empresa]["Columna"]

                #se inicia el copiado de la información desde el origen, basado en la fila inicial y final del origen
                for fila_copia in range(fila_inicial_origen, fila_final_origen + 1):

                    celda_origen = source_ws.cell(row=fila_copia, column=columna_origen)
                    celda_destino = target_ws.cell(row=fila_destino, column=columna_destino)
                    celda_destino_anterior = target_ws.cell(row=fila_destino, column=columna_destino - 1)

                    #se copia la información del archivo origen al destino
                    celda_destino.value = celda_origen.value

                    #se copian los estilos de la celda del archivo origen al destino
                    celda_destino.font = copy(celda_destino_anterior.font)
                    celda_destino.border = copy(celda_destino_anterior.border)
                    celda_destino.fill = copy(celda_destino_anterior.fill)
                    celda_destino.number_format = celda_destino_anterior.number_format
                    celda_destino.protection = copy(celda_destino_anterior.protection)
                    celda_destino.alignment = copy(celda_destino_anterior.alignment)
                    
                    #Se le suma uno a la fila destino, para que en archivo destino copie en la siguiente fila
                    fila_destino = fila_destino + 1

                if fila_destino > fila_maxima:
                    fila_maxima = fila_destino
            else:

                #print(f'No existe la empresa {empresa} en el diccionario de empresas del kit')
                fila_maxima = fila_maxima + 3


                fila_origen_origen = dic_empresas[empresa]['Filas']['inicial']
                fila_origen_final = dic_empresas[empresa]['Filas']['final']
                #columna_descripcion_origen 
                #columna_descripcion_kit
                #fila_maxima

                for fila_nuevo_dato in range(fila_origen_origen, fila_origen_final + 1):

                    celda_descripcion_origen = source_ws.cell(row=fila_nuevo_dato, column=columna_descripcion_origen)
                    celda_descripcion_destino = target_ws.cell(row=fila_maxima, column=columna_descripcion_kit)
                    celda_descripcion_destino_anterior = target_ws.cell(row=fila_maxima, column=columna_descripcion_kit - 1)

                    celda_descripcion_destino.value = celda_descripcion_origen.value

                    celda_descripcion_destino.font = copy(celda_descripcion_destino_anterior.font)
                    celda_descripcion_destino.border = copy(celda_descripcion_destino_anterior.border)
                    celda_descripcion_destino.fill = copy(celda_descripcion_destino_anterior.fill)
                    celda_descripcion_destino.number_format = celda_descripcion_destino_anterior.number_format
                    celda_descripcion_destino.protection = copy(celda_descripcion_destino_anterior.protection)
                    celda_descripcion_destino.alignment = copy(celda_descripcion_destino_anterior.alignment)

                    celda_origen = source_ws.cell(row=fila_nuevo_dato, column=columna_origen)
                    celda_destino = target_ws.cell(row=fila_maxima, column=columna_destino)
                    celda_destino_anterior = target_ws.cell(row=fila_maxima, column=columna_destino - 1)

                    celda_destino.value = celda_origen.value

                    celda_destino.font = copy(celda_destino_anterior.font)
                    celda_destino.border = copy(celda_destino_anterior.border)
                    celda_destino.fill = copy(celda_destino_anterior.fill)
                    celda_destino.number_format = celda_destino_anterior.number_format
                    celda_destino.protection = copy(celda_destino_anterior.protection)
                    celda_destino.alignment = copy(celda_destino_anterior.alignment)

                    fila_maxima = fila_maxima + 1

        with open(archivo_destino, 'wb') as f:
            target_wb.save(f)

except Exception as ex:
    print('Task Fail--> ### ' + str(ex) + ' ###')
