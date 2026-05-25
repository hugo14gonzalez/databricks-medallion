# Databricks notebook source
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
import json
from pyspark.sql.functions import col

# COMMAND ----------

# ruta_origen = '/Volumes/test_data/base/4q24/Deuda/01 Deuda_Dic2024.xlsx'
# hoja_origen = 'PerfilD'

# ruta_destino = '/Volumes/test_data/base/4q24/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx'
# hoja_destino = 'Perfil deuda'

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .filter(
        (col('Carpeta') == 'Deuda') &
        (col('HojaOrigen') == 'PerfilD')

    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

archivo_destino = dbutils.widgets.get('ruta_archivo_destino')

#dbutils.fs.cp(archivo_destino, "/Workspace/Users/itco_e_jacevedo@intercolombia.com/Procesamiento/Generador Kit/prueba_kit.xlsx")

# full_ano = dbutils.widgets.get('Anio')
# trimestre = dbutils.widgets.get('Trimestre')
# ano = int(str(full_ano)[2:])

nombre_archivo_final = 'prueba_all_kit.xlsx'


# COMMAND ----------

# Cargar los libros de Excel
source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
target_wb = load_workbook(nombre_archivo_final, data_only=True)

# Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
source_ws = source_wb[hojas_origen[0]]
target_ws = target_wb[hojas_destino[0]]

# COMMAND ----------

for rng in list(target_ws.merged_cells.ranges):
    target_ws.unmerge_cells(str(rng))

# COMMAND ----------

valor_encontrado = False
columna_kit = 1
dic_pais_origen = {}
dic_pais_destino = {}


for fila in range(1, source_ws.max_row):
    for columna in range(1,4):
        valor = source_ws.cell(row=fila, column=columna).value

        if valor == 'PERFIL DE VENCIMIENTO POR AÑOS':
            valor_encontrado = True
            break
    
    if valor_encontrado:
        
        fila_encabezado = fila
        data_columna = columna + 3

        for fila_kit in range(2, target_ws.max_row):
            
            valor_kit = str(target_ws.cell(row=fila_kit, column=columna_kit).value)

            if valor_kit != 'None':

                for sub_fila in range(fila, source_ws.max_row):
                    sub_valor = str(source_ws.cell(row=sub_fila, column=columna + 1).value)

                    if valor_kit == sub_valor:
                        #print(f'Se encontró {valor_kit} en la fila {sub_fila}')

                        dic_pais_origen[sub_valor] = { 'FilaData': sub_fila}

                        dic_pais_destino[valor_kit] = { 'FilaData': fila_kit}

                for columna_encabezado_destino in range(2, 10):
                    valor_encabezado_destino = str(target_ws.cell(row=fila_kit, column=columna_encabezado_destino).value)

                    if valor_encabezado_destino == str(source_ws.cell(row=fila, column=columna_encabezado_destino).value):
                        break
        break

print(f'Fila encabezado origen: {fila_encabezado} -- columna info origen {data_columna}')
print(dic_pais_origen)


print(f'columna info destino {columna_encabezado_destino}')
print(dic_pais_destino)

        

# COMMAND ----------


for pais in dic_pais_origen:
    for fila_data_copiar in range(0,2):
        columna_vacia = 0
        columna_destino = columna_encabezado_destino

        for columna_data_copiar in range(data_columna, source_ws.max_column + 1):
            celda_origen = source_ws.cell(row=dic_pais_origen[pais]['FilaData'] + fila_data_copiar, column=columna_data_copiar)
            celda_destino = target_ws.cell(row=dic_pais_destino[pais]['FilaData'] + fila_data_copiar, column= columna_destino)

            if celda_origen.value == None:
                columna_vacia = columna_vacia + 1
            else:
                columna_vacia = 0
                celda_destino.value = celda_origen.value
            
            columna_destino = columna_destino + 1
            
            if columna_vacia >= 3:
                break

# COMMAND ----------

#Guardar los cambios en el archivo destino
with open(nombre_archivo_final, 'wb') as f:
    target_wb.save(f)
