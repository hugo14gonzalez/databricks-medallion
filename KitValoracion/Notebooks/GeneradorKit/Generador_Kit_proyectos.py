# Databricks notebook source
import json
import os
import unicodedata
import shutil
from openpyxl import load_workbook
from copy import copy 
from datetime import datetime
from openpyxl.styles import *
from pyspark.sql.functions import col, regexp_replace, lower, udf
from pyspark.sql.types import StringType

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')

df_carpetas = df_carpetas\
    .withColumn(
        'Carpeta',
        normalizar_udf(col('Carpeta'))
    )\
    .filter(
        (col('Carpeta') == 'proyectos') 

    )

rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

#archivo_kit_origen = dbutils.widgets.get('ruta_kit_origen')
archivo_destino = dbutils.widgets.get('ruta_archivo_destino')
directorio, nombre_archivo = os.path.split(archivo_destino)
temp_path = 'tmp/'
full_temp_path = temp_path + nombre_archivo
shutil.copy(archivo_destino,full_temp_path)

#dbutils.fs.cp(archivo_destino, "/Workspace/Users/itco_e_jacevedo@intercolombia.com/Procesamiento/Generador Kit/prueba_kit.xlsx")

full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
# ano = int(str(full_ano)[2:])


# COMMAND ----------

# Cargar los libros de Excel
source_wb = load_workbook(rutaArchivos[0][5:], data_only=True)
target_wb = load_workbook(full_temp_path, data_only=True)

# Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
source_ws = source_wb[hojas_origen[0]]
target_ws = target_wb[hojas_destino[0]]

# COMMAND ----------

if trimestre == '1':
    nombre_mes = 'Marzo'
elif trimestre == '2':
    nombre_mes = 'Junio'
elif trimestre == '3':
    nombre_mes = 'Septiembre'
else:
    nombre_mes = 'Diciembre'


# COMMAND ----------

kit_encabezado_encontrado = False

for fila_kit in range(2, target_ws.max_row + 1):
    for columna_kit in range(1, target_ws.max_column + 1):
        valor_kit = target_ws.cell(row=fila_kit, column=columna_kit).value

        if valor_kit == 'Filial':
            kit_encabezado_encontrado = True
            break
    if kit_encabezado_encontrado:
        break


print(f'Se encontró la fila: {fila_kit}, columna: {columna_kit}')

# COMMAND ----------

dic_pais = {}

for fila in range(2, source_ws.max_row + 1):

    valor_año_reporte = source_ws.cell(row=fila, column=1).value
    valor_mes_reporte = source_ws.cell(row=fila, column=2).value
    valor_tipo_1 = source_ws.cell(row=fila, column=6).value
    valor_nombre_proyecto = source_ws.cell(row=fila, column=14).value

    

    if str(valor_año_reporte) == str(full_ano) and valor_mes_reporte == nombre_mes and valor_tipo_1 != 'Renovación' and valor_nombre_proyecto != 'Proyectos de refuerzos en ejecución' :

        print(f'{valor_año_reporte} -- {valor_mes_reporte} -- {valor_tipo_1} -- {valor_nombre_proyecto}')
        print()

        
        Pais = source_ws.cell(row=fila, column=3).value

        if Pais not in dic_pais:
            dic_pais[Pais] = {}

        Empresa = source_ws.cell(row=fila, column=4).value
        
        if 'REP' == Empresa:
            Empresa = 'CTM'
        
        if Empresa not in dic_pais[Pais]:
            dic_pais[Pais][Empresa] = {}

        Tipo_contol = source_ws.cell(row=fila, column=6).value

        if Tipo_contol not in dic_pais[Pais][Empresa]:
            dic_pais[Pais][Empresa][Tipo_contol] = []

        fecha_energizacion_RI = source_ws.cell(row=fila, column=12).value

        # Extraes el año
        anio = fecha_energizacion_RI.year

        # Calculas el trimestre
        trimestre = (fecha_energizacion_RI.month - 1)//3 + 1

        nombre_proyecto = source_ws.cell(row=fila, column=14).value

        comentarios = source_ws.cell(row=fila, column=23).value

        dic_pais[Pais][Empresa][Tipo_contol].append([nombre_proyecto,anio, trimestre, comentarios])

        

        #print(f' Pais: {Pais}, Empresa: {Empresa}, Tipo_contol: {Tipo_contol},Fecha: {fecha_energizacion_RI}, Año: {anio}, Trimestre: {trimestre}, comentarios: {comentarios}')

dic_pais = dict(sorted(dic_pais.items()))

#print(dic_pais)

# COMMAND ----------

for rng in list(target_ws.merged_cells.ranges):
    target_ws.unmerge_cells(str(rng))

# COMMAND ----------

for fila in range((fila_kit + 1), target_ws.max_row + 1):
    for columna in range(columna_kit, target_ws.max_column + 1):
        celda_destino = target_ws.cell(row=fila, column=columna)
        # Ahora queremos borrarlo todo:
        celda_destino.value = None  # Elimina el texto/valor
        # Para quitar el formato, asignamos estilos "vacíos"
        celda_destino.fill = PatternFill()      # Sin relleno
        celda_destino.border = Border()         # Sin bordes
        celda_destino.alignment = Alignment()   # Alineación por defecto

        celda_destino.number_format = "General" # Formato numérico por defecto

# COMMAND ----------


celda_encabezado = target_ws.cell(row=fila_kit - 1, column=columna_kit)

celda_encabezado.alignment = Alignment(horizontal="center", vertical="center")

target_ws.merge_cells(
            start_row=fila_kit - 1,
            start_column=columna_kit,
            end_row=fila_kit - 1,
            end_column=columna_kit + 6
        )


celda_encabezado = target_ws.cell(row=fila_kit, column=columna_kit)
celda_encabezado.value = 'País'

celda_encabezado = target_ws.cell(row=fila_kit, column=columna_kit + 1)
celda_encabezado.value = 'Empresa'

celda_encabezado = target_ws.cell(row=fila_kit, column=columna_kit + 2)
celda_encabezado.value = 'Tipo de control'

celda_encabezado = target_ws.cell(row=fila_kit, column=columna_kit + 3)
celda_encabezado.value = 'Nombre Proyecto'

celda_encabezado = target_ws.cell(row=fila_kit, column=columna_kit + 4)
celda_encabezado.value = 'Año'

celda_encabezado = target_ws.cell(row=fila_kit, column=columna_kit + 5)
celda_encabezado.value = 'Trimestre'


celda_encabezado = target_ws.cell(row=fila_kit, column=columna_kit + 6)
celda_encabezado.font = copy(target_ws.cell(row=fila_kit, column=columna_kit + 1).font)
celda_encabezado.fill = copy(target_ws.cell(row=fila_kit, column=columna_kit + 1).fill)
celda_encabezado.alignment = copy(target_ws.cell(row=fila_kit, column=columna_kit + 1).alignment)
celda_encabezado.value = 'Comentarios'



# COMMAND ----------

fila = 1
border_inferior = Border(bottom=Side(border_style="thin",color='D3D3D3'))
alineamiento_celdas = Alignment(horizontal="center", vertical="center")
for pais in dic_pais:
    #print(pais)
    
    celda_destino = target_ws.cell(row=fila_kit+fila, column=columna_kit)

    celda_destino.value = pais
    celda_destino.alignment = alineamiento_celdas

    fila_inicio_combinada_pais = fila_kit+fila
    for empresa in dic_pais[pais]:

        celda_destino = target_ws.cell(row=fila_kit+fila, column=columna_kit + 1)

        celda_destino.value = empresa
        celda_destino.alignment = alineamiento_celdas

        fila_inicio_combinada_empresa = fila_kit+fila
        for tipo_control in dic_pais[pais][empresa]:
            
            celda_destino = target_ws.cell(row=fila_kit+fila, column=columna_kit + 2)

            celda_destino.value = tipo_control
            celda_destino.alignment = alineamiento_celdas

            #print(tipo_control)

            fila_inicio_combinada_Tipo_control = fila_kit+fila
            for lista_datos in dic_pais[pais][empresa][tipo_control]:
                columna_dato = columna_kit + 3
                pos_datos = 1
                for dato in lista_datos:
                    celda_destino = target_ws.cell(row=fila_kit+fila, column=columna_dato)
                    if pos_datos == 2 or pos_datos == 3 :
                        celda_destino.alignment = alineamiento_celdas
                    pos_datos += 1
                    celda_destino.value = dato
                    columna_dato += 1
                    celda_destino.border = border_inferior
                fila += 1

            target_ws.merge_cells(
                    start_row=fila_inicio_combinada_Tipo_control,
                    start_column=columna_kit + 2,
                    end_row=fila_kit+ (fila - 1),
                    end_column=columna_kit + 2
            )

            for fil in range(fila_inicio_combinada_Tipo_control,fila_kit + fila):
                cell = target_ws.cell(row=fil, column=columna_kit + 2)
                cell.border = border_inferior

            target_ws.cell(row = columna_kit + 2, column= columna_kit + 2).border = border_inferior
                #print(dato)
        target_ws.merge_cells(
            start_row=fila_inicio_combinada_empresa,
            start_column=columna_kit + 1,
            end_row=fila_kit+fila - 1,
            end_column=columna_kit + 1
        )
        #for fil in range(fila_inicio_combinada_empresa,fila_kit + fila + 1):
        #    cell = target_ws.cell(row=fil, column=columna_kit + 1)
        #    cell.border = border_inferior

        fila += 1

    target_ws.merge_cells(
            start_row=fila_inicio_combinada_pais,
            start_column=columna_kit,
            end_row=fila_kit + (fila - 2),
            end_column=columna_kit
        )
    for i in range(6):
        for fil in range(fila_inicio_combinada_pais,fila_kit + fila):
            cell = target_ws.cell(row=fil, column=columna_kit + i)
            cell.border = border_inferior

    
    #print('-' * 40)

# COMMAND ----------

#Guardar los cambios en el archivo destino
# with open(archivo_destino, 'wb') as f:
#     target_wb.save(f)

target_wb.save(full_temp_path)
shutil.copy(full_temp_path,archivo_destino)
