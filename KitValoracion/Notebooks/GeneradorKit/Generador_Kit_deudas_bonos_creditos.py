# Databricks notebook source
import json
import os
import unicodedata
import shutil
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from pyspark.sql.functions import col, regexp_replace, lower, udf
from pyspark.sql.types import StringType

# COMMAND ----------

# UDF para eliminar tildes y dejar solo letras minúsculas
def normalizar_letras(texto):
    if texto is None:
        return None
    # Quitar tildes y otros diacríticos
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    # Convertir a minúsculas
    return texto.lower()

# Registrar la UDF
normalizar_udf = udf(normalizar_letras, StringType())

# COMMAND ----------

lectura_carpetas = dbutils.widgets.get('df_base')
df_carpetas = spark.createDataFrame(json.loads(lectura_carpetas))
df_carpetas = df_carpetas\
    .withColumnRenamed('_1','Carpeta')\
    .withColumnRenamed('_2','RutaArchivo')\
    .withColumnRenamed('_3','HojaOrigen')\
    .withColumnRenamed('_4','HojaDestino')\
    .withColumn(
        'Carpeta',
        normalizar_udf(col('Carpeta'))
    )\
    .withColumn(
        'HojaOrigen',
        normalizar_udf(col('HojaOrigen'))
    )\
    .filter(
        (col('Carpeta') == 'Deuda') &
        (
            (col('HojaOrigen') == 'kitbonos') |
            (col('HojaOrigen') == 'kitcreditos')
        )
    )


rutaArchivos = [row["RutaArchivo"] for row in df_carpetas.select("RutaArchivo").collect()]
hojas_origen = [row["HojaOrigen"] for row in df_carpetas.select("HojaOrigen").collect()]
hojas_destino = [row["HojaDestino"] for row in df_carpetas.select("HojaDestino").collect()]

#archivo_kit_origen = dbutils.widgets.get('ruta_kit_origen')
archivo_destino = dbutils.widgets.get('ruta_archivo_destino')
directorio, nombre_archivo = os.path.split(archivo_destino)
temp_path = 'tmp/'
full_temp_path = temp_path + nombre_archivo
shutil.copy(archivo_destino, full_temp_path)

full_ano = dbutils.widgets.get('Anio')
trimestre = dbutils.widgets.get('Trimestre')
ano = int(str(full_ano)[2:])

# COMMAND ----------

for i in range(len(rutaArchivos)):
    print(f'Archivo {i+1}/{len(rutaArchivos)}: {rutaArchivos[i]}...')
    print(f'Hoja: {hojas_origen[i]}...')
    print('=' * 80)

    # Cargar los libros de Excel
    source_wb = load_workbook(rutaArchivos[i][5:], data_only=True)
    target_wb = load_workbook(full_temp_path, data_only=True)

    # Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
    source_ws = source_wb[hojas_origen[i]]
    target_ws = target_wb[hojas_destino[i]]

    for rng in list(target_ws.merged_cells.ranges):
        target_ws.unmerge_cells(str(rng))
    
    for fila in range(1,200):
        for columna in range(1,16):
            
            celda_origen = source_ws.cell(row=fila, column=columna)
            celda_destino = target_ws.cell(row=fila, column=columna)

            # Borrar contenido
            celda_destino.value = None
            
            # Restablecer estilos
            celda_destino.font = Font()              # Fuente por defecto
            celda_destino.border = Border()          # Sin bordes
            celda_destino.fill = PatternFill()       # Sin relleno
            celda_destino.number_format = 'General'  # Formato numérico general
            celda_destino.protection = Protection()  # Protección por defecto
            celda_destino.alignment = Alignment()    # Alineación por defecto



            celda_destino.font = copy(celda_origen.font)
            celda_destino.fill = copy(celda_origen.fill)
            celda_destino.number_format = celda_origen.number_format
            celda_destino.alignment = copy(celda_origen.alignment)

            celda_destino.value = celda_origen.value
    
    # with open(archivo_destino, 'wb') as f:
    #     wb_destino.save(f)

    wb_destino.save(full_temp_path)
    shutil.copy(full_temp_path,archivo_destino)
