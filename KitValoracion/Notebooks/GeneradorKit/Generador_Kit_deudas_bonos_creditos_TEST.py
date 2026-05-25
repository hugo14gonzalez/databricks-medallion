# Databricks notebook source
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from pyspark.sql.functions import col
import json

# COMMAND ----------

ruta_origen = ['abfs:/Volumes/test_data/base/4q23/Deuda/01 Deuda_Jun2024 - VF_3Q_2024.xlsx']
hojas_origen = ['KitBonos', 'KitCreditos']

ruta_destino = ['abfs:/Volumes/test_data/base/4q24/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx']
hojas_destino = ['Deuda consolidada bonos', 'Deuda consolidada créditos']

# COMMAND ----------

nombre_archivo_final = 'prueba_deuda_kit.xlsx'

# Cargar los libros de Excel
source_wb = load_workbook(ruta_origen[0][5:], data_only=True)
target_wb = load_workbook(ruta_destino[0][5:], data_only=True)

# Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
source_ws = source_wb[hojas_origen[1]]
target_ws = target_wb[hojas_destino[1]]

for rng in list(target_ws.merged_cells.ranges):
    target_ws.unmerge_cells(str(rng))

for fila in range(1,200):
    for columna in range(1,16):
        
        celda_origen = source_ws.cell(row=fila, column=columna)
        celda_destino = target_ws.cell(row=fila, column=columna)

        # Borrar contenido
        celda_destino.value = None
        
        # Restablecer estilos
        celda_destino.font = Font()              # Fuente por defecto
        celda_destino.border = Border()          # Sin bordes
        celda_destino.fill = PatternFill()       # Sin relleno
        celda_destino.number_format = 'General'  # Formato numérico general
        celda_destino.protection = Protection()  # Protección por defecto
        celda_destino.alignment = Alignment()    # Alineación por defecto
        
        celda_destino.font = copy(celda_origen.font)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.number_format = celda_origen.number_format
        celda_destino.alignment = copy(celda_origen.alignment)

        celda_destino.value = celda_origen.value



# COMMAND ----------

with open(nombre_archivo_final, 'wb') as f:
    target_wb.save(f)
