# Databricks notebook source
from openpyxl import load_workbook
from copy import copy
import json
from pyspark.sql.functions import col

# COMMAND ----------

ruta_origen = ['absf:/Volumes/test_data/base/4q24/Participaciones Empresas/RP_Participacion Directa & Efectiva1224.xlsm']
hojas_origen = ['Reporte_SP']

archivo_destino = ['absf:/Volumes/test_data/base/4q24/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx']
hojas_destino = ['Participación en compañías']

full_ano = 2024
trimestre = 3
ano = int(str(full_ano)[2:])

nombre_archivo_final = 'prueba_participacion_kit.xlsx'

# COMMAND ----------

# Cargar los libros de Excel
source_wb = load_workbook(ruta_origen[0][5:], data_only=True)
target_wb = load_workbook(archivo_destino[0][5:], data_only=True)

# Seleccionar las hojas correspondientes (ajusta el nombre de la hoja según tu archivo)
source_ws = source_wb[hojas_origen[0]]
target_ws = target_wb[hojas_destino[0]]

# COMMAND ----------

lista_ubicacion_cuadros_origen = []
columna_final = 0

for fila in range(1, source_ws.max_row + 1):
    for columna in range(1, source_ws.max_column + 1):
        valor = str(source_ws.cell(row=fila, column=columna).value)
        if valor == 'Tipo de \nControl':
            lista_ubicacion_cuadros_origen.append([fila, columna])
        
        elif valor == "#" or valor == 'SocIndirecta':
            lista_ubicacion_cuadros_origen[columna_final].append(columna)
            columna_final += 1


# lista_ubicacion_origen = [lista_ubicacion_cuadros_origen[0]]
# lista_ubicacion_origen.append([lista_ubicacion_cuadros_origen[-1]])


print(lista_ubicacion_cuadros_origen)

# COMMAND ----------

lista_ubicacion_cuadros_destino = []
columna_final = 0

for fila in range(1, target_ws.max_row + 1):
    for columna in range(1, target_ws.max_column + 1):
        valor = str(target_ws.cell(row=fila, column=columna).value)
        #print(f'{fila}, {columna}) {valor}')
        if valor == 'Tipo de \nControl':
            lista_ubicacion_cuadros_destino.append([fila, columna])
        
        elif valor == "#" or valor == 'SocIndirecta':
            
            lista_ubicacion_cuadros_destino[columna_final].append(columna)
            columna_final += 1

# lista_ubicacion_destino = [lista_ubicacion_cuadros_destino[0]]
#lista_ubicacion_destino = lista_ubicacion_cuadros_destino[:-1]

print(lista_ubicacion_cuadros_destino)
#print(lista_ubicacion_destino)

# COMMAND ----------

for rng in list(target_ws.merged_cells.ranges):
    target_ws.unmerge_cells(str(rng))

# COMMAND ----------

f = ''
fila_destino = lista_ubicacion_cuadros_destino[0][0]
columna_destino_fin = lista_ubicacion_cuadros_destino[0][2]

for fila in range(lista_ubicacion_cuadros_origen[0][0], 10):
    columna_destino = lista_ubicacion_cuadros_destino[0][1]
    for columna in range(lista_ubicacion_cuadros_origen[0][1], lista_ubicacion_cuadros_origen[0][2] + 1):
        celda_origen = source_ws.cell(row=fila, column=columna)
        celda_destino = target_ws.cell(row=fila_destino, column=columna_destino)

        celda_destino.value = str(celda_origen.value)

        if celda_destino.value == 'None':
             target_ws.merge_cells(
                        start_row=fila_destino,
                        start_column=columna_destino - 1,
                        end_row=fila_destino,
                        end_column=columna_destino
                    )

        columna_destino += 1
        f = f + ' | ' + celda_destino.value

    fila_destino += 1
    f = f + '\n'

print(f)

# COMMAND ----------

for fila in range(lista_ubicacion_cuadros_origen[1][0],source_ws.max_row + 1):
    for columna in range(lista_ubicacion_cuadros_origen[1][1], lista_ubicacion_cuadros_origen[1][2] + 1):
        celda_origen = source_ws.cell(row=fila, column=columna)

        

        f = f + ' | ' + str(celda_origen.value)
    
    f = f + '\n'


print(f)
        

# COMMAND ----------

# #copiar primer cuadro
# fila_vacia = False
# fila_sin_datos = 1
# salto = False

# for fila in range(1, source_ws.max_row + 1):
#     columna_vacia = 0
#     for columna in range(1, 12):

#         celda_origen = source_ws.cell(row=fila, column=columna)
#         valor = str(celda_origen.value)

#         if valor != 'None':
#             fila_vacia = False
#             fila_sin_datos = 0

#             if salto:
#                 celda_destino = target_ws.cell(row=fila, column= columna)
                
#             else:
#                 celda_destino = target_ws.cell(row=fila + 3, column= columna)

#             celda_destino.font = copy(celda_origen.font)
#             celda_destino.value = valor  
#         else:
            

#             # if salto:
#             #     if columna - 1 > 0:
#             #         celda_anterior = target_ws.cell(row=fila + 1, column= columna - 1)

#             #         if str(celda_anterior.value) != 'None':

#             #             target_ws.merge_cells(
#             #                 start_row=fila + 1,
#             #                 start_column=columna - 1,
#             #                 end_row=fila + 1,
#             #                 end_column=columna
#             #             )
            
#             columna_vacia += 1
        
#         # print(f'fila: {fila}, columna: {columna}')
#         # print(valor)
        
#         if columna_vacia >= 11:
#             fila_vacia = True
            
#             break 
    
#     if fila_vacia:
#         fila_sin_datos += 1

#     if fila_sin_datos >= 3:
#         salto = True
    
#     # print(fila_sin_datos)
#     # print('*' * 40)
    


# COMMAND ----------

#Guardar los cambios en el archivo destino
with open(f'{nombre_archivo_final}', 'wb') as f:
    target_wb.save(f)
