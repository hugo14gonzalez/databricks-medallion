# Databricks notebook source
from openpyxl import load_workbook
import json
import pandas as pd

# COMMAND ----------

success = True
result = ''

try:
    lectura_carpetas = dbutils.widgets.get('df_listado_archivos')
    df_inicial = spark.createDataFrame(json.loads(lectura_carpetas))
    df_inicial = df_inicial\
        .withColumnRenamed('_1','Carpeta')\
        .withColumnRenamed('_2','RutaArchivos')\
        .withColumnRenamed('_3','HojaOrigen')\
        .withColumnRenamed('_4','HojaDestino')\


    rutas = [row["RutaArchivos"] for row in df_inicial.select("RutaArchivos").collect()]
    hojas = [row["HojaOrigen"] for row in df_inicial.select("HojaOrigen").collect()]

    for posicion in range(len(rutas)):
        ruta_archivo = rutas[posicion][5:]

        #hugo
        print(f'Validando: "{ruta_archivo}"...')

        if '.xlsx' in ruta_archivo:
            wb = load_workbook(ruta_archivo, data_only=True)
            hojas_catalogo = wb.sheetnames
            if hojas[posicion] in hojas_catalogo:
                #print(f'La hoja "{hojas[posicion]}" existe en el archivo "{ruta_archivo}".')
                continue
            else:
                result = result + f'La hoja "{hojas[posicion]}" no existe en el archivo "{ruta_archivo}".\n'
                success = False

    if success:
        result = 'Success'
        dbutils.jobs.taskValues.set(key = 'df_listado_archivos', value = df_inicial.collect())
    else:
        #hugo
        print(f'result: {result}')
except Exception as ex:
    success = False
    result = 'Task Fail--> ### ' + str(ex) + ' ###'
    print(f'Resul: {result}')

dbutils.jobs.taskValues.set( key = "resultado", value = result)
dbutils.jobs.taskValues.set( key = 'estado', value = success)
