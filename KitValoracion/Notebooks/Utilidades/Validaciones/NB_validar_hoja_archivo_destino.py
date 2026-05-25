# Databricks notebook source
from openpyxl import load_workbook
import json
import pandas as pd

# COMMAND ----------

ruta_origen = dbutils.widgets.get('ruta_excel_base')
ruta_destino = dbutils.widgets.get('ruta_kit_origen')

# COMMAND ----------

success = True
result = 'Success'

#ruta_origen = '/Workspace/Users/itco_e_jacevedo@intercolombia.com/NombreCarpetas_test.xlsx'
#ruta_destino = '/Volumes/test_data/base/4q23/Consolidado y separados ISA/kit_Inversionistas ISA_Sp.xlsx' 
try:
    wb_origen = load_workbook(ruta_origen, data_only=True)
    primera_hoja = wb_origen['Hoja1']
    hojas_origen = [row[2].value for row in primera_hoja.iter_rows(min_row=1, max_row=primera_hoja.max_row)]

    wb_destino = load_workbook(ruta_destino, data_only=True)
    hojas_destino = wb_destino.sheetnames

    missing_values = list(set(hojas_origen) - set(hojas_destino))
    # Display the missing values
    if len(missing_values) > 0:
        print(missing_values)

        success = False
        result = 'Task Fail--> ### ' + 'las hojas: ' + '\n' + str(missing_values) + '\n' + 'están configuradas en el archivo base, pero no están en el archivo destino' + ' ###'
except Exception as ex:
    success = False
    result = 'Task Fail--> ### ' + str(ex) + ' ###'
    print(f'Resul: {result}')

dbutils.jobs.taskValues.set( key = "resultado", value = result)
dbutils.jobs.taskValues.set( key = 'estado', value = success)
